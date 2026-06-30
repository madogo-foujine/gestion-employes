from __future__ import annotations

import calendar
import csv
import datetime as dt
import hashlib
import json
import logging
import os
import secrets
import shutil
import webbrowser
from pathlib import Path

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import ttkbootstrap as tb

try:
    from ttkbootstrap.widgets import ToastNotification, ToolTip
    HAS_TB_EXTRAS = True
except Exception:  # noqa: BLE001
    try:
        from ttkbootstrap.toast import ToastNotification
        from ttkbootstrap.tooltip import ToolTip
        HAS_TB_EXTRAS = True
    except Exception:  # noqa: BLE001
        HAS_TB_EXTRAS = False

THEME_LIGHT = "flatly"
THEME_DARK = "darkly"
APP_VERSION = "1.1"

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    raise SystemExit(
        "خاصك تثبت المكتبة openpyxl. كتب فالـTerminal:\n\n"
        "    pip install openpyxl\n"
    )

try:
    from PIL import Image, ImageDraw, ImageTk
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

try:
    from fpdf import FPDF
    HAS_FPDF = True
except ImportError:
    HAS_FPDF = False


APP_TITLE = "Gestion des Employés & Paie — تسيير الخدامة والأجور"
COMPANY_NAME = "Ma Société"
LOGO_PATH = ""
SIGN_PATH = ""
DEFAULT_FILE = Path.home() / "employes.xlsx"
CONFIG_PATH = Path.home() / ".employee_manager.json"
LOG_PATH = Path.home() / ".employee_manager.log"
SHEET_NAME = "Employes"

logging.basicConfig(
    filename=str(LOG_PATH), level=logging.WARNING,
    format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("employee_manager")
JOURS_OUVRABLES = 26
LEAVE_PER_MONTH = 1.5
STD_START = "09:00"                    # ساعة بداية العمل (لحساب التأخير)
STD_DAILY_HOURS = 8.0                  # ساعات العمل القياسية فاليوم
OVERTIME_RATE = 1.25                   # معامل الساعات الإضافية

PLAFOND_CNSS = 6000.0
TAUX_CNSS = 0.0448
TAUX_AMO = 0.0226
SEUIL_FP = 6500.0
PLAFOND_FP_MENSUEL = 2916.67
DEDUCTION_CHARGE = 30.0
IR_BRACKETS = [
    (3333.33, 0.00),
    (5000.00, 0.10),
    (6666.67, 0.20),
    (8333.33, 0.30),
    (15000.00, 0.34),
    (float("inf"), 0.37),
]

PALETTES = {
    "light": {
        "brand":   "#0f3d2e", "brand2":  "#15803d", "accent":  "#16a34a",
        "danger":  "#dc2626", "info":    "#2563eb", "bg":      "#eef2f5",
        "surface": "#ffffff", "border":  "#dce3ea", "text":    "#1e293b",
        "muted":   "#64748b", "row_alt": "#f6f9f7", "sel":     "#16a34a",
        "net_bg":  "#eafaf0", "ded_bg":  "#fef2f2", "ded_fg":  "#b91c1c",
    },
    "dark": {
        "brand":   "#0b2a20", "brand2":  "#16a34a", "accent":  "#22c55e",
        "danger":  "#ef4444", "info":    "#3b82f6", "bg":      "#0f172a",
        "surface": "#1e293b", "border":  "#334155", "text":    "#e2e8f0",
        "muted":   "#94a3b8", "row_alt": "#243244", "sel":     "#16a34a",
        "net_bg":  "#10331f", "ded_bg":  "#3a1a1a", "ded_fg":  "#fca5a5",
    },
}
COL = dict(PALETTES["light"])
AVATAR_COLORS = ["#16a34a", "#2563eb", "#9333ea", "#db2777",
                 "#ea580c", "#0891b2", "#ca8a04", "#be123c"]
FONT = "Segoe UI"


def load_config() -> dict:
    try:
        return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return {}


def save_config(cfg: dict) -> None:
    try:
        CONFIG_PATH.write_text(json.dumps(cfg, indent=2), encoding="utf-8")
    except OSError:
        log.exception("Echec de l'ecriture de la configuration")


def hash_pw(password: str, salt: str) -> str:
    return hashlib.scrypt(
        password.encode("utf-8"), salt=bytes.fromhex(salt),
        n=16384, r=8, p=1, dklen=32, maxmem=67108864).hex()


def make_password(password: str) -> dict:
    salt = secrets.token_hex(16)
    return {"pw_algo": "scrypt", "pw_salt": salt, "pw_hash": hash_pw(password, salt)}


def verify_hash(password: str, salt: str, stored: str, algo: str = "scrypt") -> bool:
    try:
        if algo == "scrypt":
            return secrets.compare_digest(hash_pw(password, salt), stored)
        legacy = hashlib.sha256((salt + password).encode("utf-8")).hexdigest()
        return secrets.compare_digest(legacy, stored)
    except (ValueError, TypeError):
        log.exception("Echec de verification du mot de passe")
        return False


def verify_pw(password: str, cfg: dict) -> bool:
    stored = cfg.get("pw_hash")
    if not stored:
        return True
    return verify_hash(password, cfg.get("pw_salt", ""), stored,
                       cfg.get("pw_algo", "sha256"))


SETTING_DEFAULTS = {
    "company_name": "Ma Société", "logo_path": "", "signature_path": "",
    "jours_ouvrables": 26, "leave_per_month": 1.5, "plafond_cnss": 6000.0,
    "taux_cnss": 0.0448, "taux_amo": 0.0226, "seuil_fp": 6500.0,
    "plafond_fp": 2916.67, "deduction_charge": 30.0,
}


def _num(cfg, key, default, lo, hi, cast=float):
    """قراءة قيمة رقمية من الإعدادات مع التحقق من المجال (وإلا القيمة الافتراضية)."""
    try:
        val = cast(cfg.get(key, default))
    except (TypeError, ValueError):
        log.warning("Parametre invalide %r=%r, valeur par defaut utilisee",
                    key, cfg.get(key))
        return default
    if not (lo <= val <= hi):
        log.warning("Parametre %r hors limites (%r), valeur par defaut utilisee",
                    key, val)
        return default
    return val


def apply_config_settings(cfg: dict) -> None:
    global COMPANY_NAME, LOGO_PATH, SIGN_PATH, JOURS_OUVRABLES, LEAVE_PER_MONTH
    global PLAFOND_CNSS, TAUX_CNSS, TAUX_AMO, SEUIL_FP
    global PLAFOND_FP_MENSUEL, DEDUCTION_CHARGE, IR_BRACKETS
    COMPANY_NAME = str(cfg.get("company_name") or "Ma Société")[:120]
    LOGO_PATH = str(cfg.get("logo_path", "") or "")
    SIGN_PATH = str(cfg.get("signature_path", "") or "")
    JOURS_OUVRABLES = _num(cfg, "jours_ouvrables", 26, 1, 31, int)
    LEAVE_PER_MONTH = _num(cfg, "leave_per_month", 1.5, 0, 31)
    PLAFOND_CNSS = _num(cfg, "plafond_cnss", 6000.0, 0, 1_000_000)
    TAUX_CNSS = _num(cfg, "taux_cnss", 0.0448, 0, 1)
    TAUX_AMO = _num(cfg, "taux_amo", 0.0226, 0, 1)
    SEUIL_FP = _num(cfg, "seuil_fp", 6500.0, 0, 1_000_000)
    PLAFOND_FP_MENSUEL = _num(cfg, "plafond_fp", 2916.67, 0, 1_000_000)
    DEDUCTION_CHARGE = _num(cfg, "deduction_charge", 30.0, 0, 100_000)
    br = cfg.get("ir_brackets")
    if br:
        try:
            brackets = [
                (float("inf") if c in (None, "") else float(c), float(r))
                for c, r in br]
            if brackets and all(0 <= r <= 1 for _, r in brackets):
                IR_BRACKETS = brackets
            else:
                log.warning("Bareme IR invalide, valeurs par defaut conservees")
        except (TypeError, ValueError):
            log.warning("Bareme IR illisible, valeurs par defaut conservees")


FIELDS = [
    ("id",              "ID",                 "base",   "readonly"),
    ("photo",           "Photo",              "meta",   "hidden"),
    ("archive",         "Archivé",            "meta",   "hidden"),
    ("nom",             "Nom complet",        "base",   "text"),
    ("cin",             "CIN",                "base",   "text"),
    ("poste",           "Poste",              "base",   "text"),
    ("salaire_base",    "Salaire de base",    "base",   "number"),

    ("telephone",       "Téléphone",          "contact", "text"),
    ("email",           "Email",              "contact", "text"),
    ("adresse",         "Adresse",            "contact", "text"),

    ("date_naissance",  "Date de naissance",  "admin",  "date"),
    ("date_embauche",   "Date d'embauche",    "admin",  "date"),
    ("type_contrat",    "Type de contrat",    "admin",  "combo"),
    ("date_fin_contrat", "Fin de contrat",    "admin",  "date"),
    ("cnss",            "N° CNSS",            "admin",  "text"),
    ("personnes_charge", "Personnes à charge", "admin", "number"),

    ("primes",          "Primes",             "fin",    "number"),
    ("retenues",        "Autres retenues",    "fin",    "number"),
    ("retenue_avance",  "Retenue avance",     "fin",    "number"),
    ("jours_absence",   "Jours d'absence",    "fin",    "number"),

    ("salaire_brut",    "Salaire brut",       "calc",   "readonly"),
    ("ret_cnss",        "CNSS (4.48%)",       "calc",   "readonly"),
    ("ret_amo",         "AMO (2.26%)",        "calc",   "readonly"),
    ("ret_ir",          "IR",                 "calc",   "readonly"),
    ("anciennete",      "Ancienneté",         "calc",   "readonly"),
    ("salaire_net",     "Salaire net",        "calc",   "readonly"),
]

FIELD_BY_KEY = {f[0]: f for f in FIELDS}
HEADERS = [f[1] for f in FIELDS]
HEADER_TO_KEY = {f[1]: f[0] for f in FIELDS}
CALC_KEYS = [k for k, _, g, _ in FIELDS if g == "calc"]

GROUPS = {
    "base":    ("Informations de base", "👤"),
    "contact": ("Contact", "📞"),
    "admin":   ("Administratif", "🗂"),
    "fin":     ("Salaire & éléments variables", "💰"),
}

CONTRAT_OPTIONS = ["CDI", "CDD", "ANAPEC", "Stage", "Intérim", "Autre"]


def to_float(value) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(str(value).replace(",", ".").replace(" ", ""))
    except (TypeError, ValueError):
        return 0.0


def fmt_money(value) -> str:
    return f"{to_float(value):,.2f}".replace(",", " ") + " DH"


def parse_time(value):
    """'HH:MM' -> minutes depuis minuit, ou None."""
    if not value:
        return None
    txt = str(value).strip().replace("h", ":").replace("H", ":")
    parts = txt.split(":")
    try:
        h = int(parts[0])
        m = int(parts[1]) if len(parts) > 1 and parts[1] != "" else 0
    except (ValueError, IndexError):
        return None
    if 0 <= h <= 23 and 0 <= m <= 59:
        return h * 60 + m
    return None


def compute_day_hours(t_in, t_out):
    """كيرجع (heures, retard_min, heures_sup) ليوم واحد."""
    mi, mo = parse_time(t_in), parse_time(t_out)
    if mi is None or mo is None or mo <= mi:
        return 0.0, 0, 0.0
    hours = round((mo - mi) / 60, 2)
    start = parse_time(STD_START) or 0
    retard = max(0, mi - start)
    sup = round(max(0.0, hours - STD_DAILY_HOURS), 2)
    return hours, retard, sup


def parse_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value).strip().split(" ")[0]
    for sep in ("-", "/"):
        parts = text.split(sep)
        if len(parts) == 3:
            try:
                a, b, c = (int(p) for p in parts)
            except ValueError:
                continue
            if a > 31:
                return dt.date(a, b, c)
            return dt.date(c, b, a)
    return None


def compute_anciennete(date_embauche) -> str:
    d = parse_date(date_embauche)
    if not d:
        return ""
    today = dt.date.today()
    months = (today.year - d.year) * 12 + (today.month - d.month)
    if today.day < d.day:
        months -= 1
    months = max(months, 0)
    years, rem = divmod(months, 12)
    parts = []
    if years:
        parts.append(f"{years} an" + ("s" if years > 1 else ""))
    parts.append(f"{rem} mois")
    return " ".join(parts)


def compute_ir_brut(sni: float) -> float:
    ir, lower = 0.0, 0.0
    for cap, rate in IR_BRACKETS:
        if sni > lower:
            ir += (min(sni, cap) - lower) * rate
            lower = cap
        else:
            break
    return round(ir, 2)


def compute_payroll(rec: dict) -> dict:
    base = to_float(rec.get("salaire_base"))
    primes = to_float(rec.get("primes"))
    autres = to_float(rec.get("retenues"))
    avance = to_float(rec.get("retenue_avance"))
    jours_abs = to_float(rec.get("jours_absence"))
    charges = int(to_float(rec.get("personnes_charge")))

    brut = base + primes
    daily = base / JOURS_OUVRABLES if JOURS_OUVRABLES else 0
    ded_absence = round(daily * jours_abs, 2)

    cnss = round(min(brut, PLAFOND_CNSS) * TAUX_CNSS, 2)
    amo = round(brut * TAUX_AMO, 2)
    taux_fp = 0.35 if brut <= SEUIL_FP else 0.25
    fp = round(min(brut * taux_fp, PLAFOND_FP_MENSUEL), 2)
    sni = max(0.0, brut - fp - cnss - amo)
    ir = max(0.0, round(compute_ir_brut(sni) - DEDUCTION_CHARGE * min(charges, 6), 2))

    total_retenues = round(cnss + amo + ir + autres + avance + ded_absence, 2)
    net = round(brut - cnss - amo - ir - autres - avance - ded_absence, 2)
    return {
        "base": round(base, 2), "primes": round(primes, 2),
        "brut": round(brut, 2), "cnss": cnss, "amo": amo, "fp": fp,
        "sni": round(sni, 2), "ir": ir, "ded_absence": ded_absence,
        "autres": round(autres, 2), "avance": round(avance, 2),
        "total_retenues": total_retenues, "net": net, "charges": charges,
    }


def compute_salaire_net(rec: dict) -> float:
    return compute_payroll(rec)["net"]


def initials(name: str) -> str:
    parts = [p for p in str(name).split() if p]
    if not parts:
        return "?"
    if len(parts) == 1:
        return parts[0][:2].upper()
    return (parts[0][0] + parts[1][0]).upper()


class ExcelStore:
    def __init__(self, path: Path):
        self.path = Path(path)
        if not self.path.exists():
            self._create_empty()

    def _create_empty(self):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADERS)
        wb.save(self.path)

    def load(self) -> list[dict]:
        wb = load_workbook(self.path, data_only=True)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header_row = [str(h).strip() if h is not None else "" for h in rows[0]]
        records = []
        for raw in rows[1:]:
            if raw is None or all(c is None or c == "" for c in raw):
                continue
            rec = {}
            for i, header in enumerate(header_row):
                key = HEADER_TO_KEY.get(header)
                if key is None:
                    continue
                val = raw[i] if i < len(raw) else None
                if isinstance(val, dt.datetime):
                    val = val.date().isoformat()
                rec[key] = "" if val is None else val
            for k in FIELD_BY_KEY:
                rec.setdefault(k, "")
            records.append(rec)
        return records

    def _backup(self, keep: int = 20):
        if not self.path.exists():
            return
        bdir = self.path.parent / "backups"
        bdir.mkdir(exist_ok=True)
        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = bdir / f"{self.path.stem}_{ts}{self.path.suffix}"
        try:
            shutil.copy2(self.path, dest)
        except OSError:
            log.exception("Echec de la sauvegarde de %s", self.path)
            return
        backups = sorted(bdir.glob(f"{self.path.stem}_*{self.path.suffix}"))
        for old in backups[:-keep]:
            try:
                old.unlink()
            except OSError:
                log.warning("Impossible de supprimer l'ancienne sauvegarde %s", old)

    def save_all(self, records: list[dict]):
        self._backup()
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADERS)
        for rec in records:
            ws.append([rec.get(HEADER_TO_KEY[h], "") for h in HEADERS])
        for col_idx, header in enumerate(HEADERS, start=1):
            letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[letter].width = max(12, len(header) + 4)
        wb.save(self.path)


POINTAGE_STATUSES = ["P", "A", "C", "R"]
POINTAGE_LABELS = {"P": "Présent", "A": "Absent", "C": "Congé", "R": "Repos"}
POINTAGE_COLORS = {"P": "#16a34a", "A": "#dc2626", "C": "#2563eb", "R": "#94a3b8"}


class PointageStore:
    def __init__(self, excel_path: Path):
        self.path = Path(excel_path).parent / "pointage.json"

    def _load(self) -> dict:
        try:
            return json.loads(self.path.read_text(encoding="utf-8"))
        except (OSError, ValueError):
            return {}

    def get(self, emp_id, mois: str) -> list:
        return self._load().get(str(emp_id), {}).get(mois, [])

    def all_for(self, emp_id) -> dict:
        return self._load().get(str(emp_id), {})

    def count_status(self, emp_id, code: str) -> int:
        return sum(months.count(code)
                   for months in self.all_for(emp_id).values())

    def set(self, emp_id, mois: str, statuses: list):
        data = self._load()
        data.setdefault(str(emp_id), {})[mois] = statuses
        try:
            self.path.write_text(json.dumps(data, ensure_ascii=False, indent=2),
                                 encoding="utf-8")
        except OSError:
            log.exception("Echec de l'ecriture du pointage")


def default_pointage(year: int, month: int) -> list:
    ndays = calendar.monthrange(year, month)[1]
    out = []
    for d in range(1, ndays + 1):
        wd = dt.date(year, month, d).weekday()
        out.append("R" if wd >= 5 else "P")
    return out


class HistoryStore:
    def __init__(self, excel_path: Path):
        self.path = Path(excel_path).parent / "history.json"

    def load(self) -> dict:
        try:
            return json.loads(self.path.read_text(encoding="utf-8"))
        except (OSError, ValueError):
            return {}

    def archive(self, mois: str, records: list, payroll_fn):
        data = self.load()
        tot = {"brut": 0.0, "cnss": 0.0, "amo": 0.0, "ir": 0.0, "net": 0.0}
        for rec in records:
            p = payroll_fn(rec)
            for k in tot:
                tot[k] += p[k]
        data[mois] = {
            "date": dt.datetime.now().isoformat(timespec="seconds"),
            "effectif": len(records),
            **{k: round(v, 2) for k, v in tot.items()},
        }
        try:
            self.path.write_text(json.dumps(data, ensure_ascii=False, indent=2),
                                 encoding="utf-8")
        except OSError:
            log.exception("Echec de l'ecriture de l'historique")

    def series(self) -> list:
        return sorted(self.load().items())


class AdvancesStore:
    def __init__(self, excel_path: Path):
        self.path = Path(excel_path).parent / "advances.json"

    def _load(self) -> dict:
        try:
            return json.loads(self.path.read_text(encoding="utf-8"))
        except (OSError, ValueError):
            return {}

    def _save(self, data: dict):
        try:
            self.path.write_text(json.dumps(data, ensure_ascii=False, indent=2),
                                 encoding="utf-8")
        except OSError:
            log.exception("Echec de l'ecriture des avances")

    def list(self, emp_id) -> list:
        return self._load().get(str(emp_id), [])

    def set(self, emp_id, advances: list):
        data = self._load()
        data[str(emp_id)] = advances
        self._save(data)

    def add(self, emp_id, montant: float, mensualite: float):
        advances = self.list(emp_id)
        advances.append({
            "date": dt.date.today().isoformat(),
            "montant": round(montant, 2),
            "mensualite": round(mensualite, 2),
            "solde": round(montant, 2),
        })
        self.set(emp_id, advances)

    def solde_total(self, emp_id) -> float:
        return round(sum(a.get("solde", 0) for a in self.list(emp_id)), 2)


class AuditStore:
    def __init__(self, excel_path: Path):
        self.path = Path(excel_path).parent / "audit.json"

    def entries(self) -> list:
        try:
            return json.loads(self.path.read_text(encoding="utf-8"))
        except (OSError, ValueError):
            return []

    def log(self, user: str, action: str, detail: str):
        data = self.entries()
        data.append({
            "time": dt.datetime.now().isoformat(timespec="seconds"),
            "user": user, "action": action, "detail": detail,
        })
        try:
            self.path.write_text(json.dumps(data, ensure_ascii=False, indent=2),
                                 encoding="utf-8")
        except OSError:
            log.exception("Echec de l'ecriture du journal d'audit")


LEAVE_TYPES = ["Annuel", "Maladie", "Sans solde"]
LEAVE_STATUS = ["En attente", "Approuvé", "Refusé"]


class LeaveStore:
    def __init__(self, excel_path: Path):
        self.path = Path(excel_path).parent / "leaves.json"

    def _load(self) -> dict:
        try:
            return json.loads(self.path.read_text(encoding="utf-8"))
        except (OSError, ValueError):
            return {}

    def _save(self, data: dict):
        try:
            self.path.write_text(json.dumps(data, ensure_ascii=False, indent=2),
                                 encoding="utf-8")
        except OSError:
            log.exception("Echec de l'ecriture des conges")

    def list(self, emp_id) -> list:
        return self._load().get(str(emp_id), [])

    def set_all(self, emp_id, requests: list):
        data = self._load()
        data[str(emp_id)] = requests
        self._save(data)

    def add(self, emp_id, typ: str, start: str, end: str, days: int):
        reqs = self.list(emp_id)
        reqs.append({
            "type": typ, "start": start, "end": end, "days": days,
            "status": "En attente",
            "demande": dt.date.today().isoformat(),
        })
        self.set_all(emp_id, reqs)

    def approved_annual_days(self, emp_id) -> int:
        return sum(int(r.get("days", 0)) for r in self.list(emp_id)
                   if r.get("type") == "Annuel" and r.get("status") == "Approuvé")


class HoursStore:
    def __init__(self, excel_path: Path):
        self.path = Path(excel_path).parent / "hours.json"

    def _load(self) -> dict:
        try:
            return json.loads(self.path.read_text(encoding="utf-8"))
        except (OSError, ValueError):
            return {}

    def get_month(self, emp_id, mois: str) -> dict:
        return self._load().get(str(emp_id), {}).get(mois, {})

    def set_day(self, emp_id, mois: str, day: int, t_in: str, t_out: str):
        data = self._load()
        emp = data.setdefault(str(emp_id), {})
        month = emp.setdefault(mois, {})
        if not t_in and not t_out:
            month.pop(str(day), None)
        else:
            month[str(day)] = {"in": t_in, "out": t_out}
        try:
            self.path.write_text(json.dumps(data, ensure_ascii=False, indent=2),
                                 encoding="utf-8")
        except OSError:
            log.exception("Echec de l'ecriture des heures")


class EmployeeApp(tb.Window):
    def __init__(self):
        cfg = load_config()
        theme = cfg.get("theme", "light")
        super().__init__(themename=THEME_DARK if theme == "dark" else THEME_LIGHT)
        self.title(APP_TITLE)
        self.geometry("1200x800")
        self.minsize(1000, 640)

        self.config_data = cfg
        apply_config_settings(self.config_data)
        self.theme = theme
        COL.clear()
        COL.update(PALETTES.get(self.theme, PALETTES["light"]))
        self.configure(bg=COL["bg"])

        self.withdraw()
        self._show_splash()

        self.store = ExcelStore(DEFAULT_FILE)
        self.pointage = PointageStore(DEFAULT_FILE)
        self.history = HistoryStore(DEFAULT_FILE)
        self.advances = AdvancesStore(DEFAULT_FILE)
        self.audit = AuditStore(DEFAULT_FILE)
        self.leaves = LeaveStore(DEFAULT_FILE)
        self.hours = HoursStore(DEFAULT_FILE)
        self.records: list[dict] = []
        self.vars: dict[str, tk.StringVar] = {}
        self.detail: dict[str, tk.StringVar] = {}
        self.current_index: int | None = None

        if not self._require_login():
            self._close_splash()
            self.destroy()
            return

        self.build_ui()
        self.reload()
        self.after(1500, self._close_splash)

    def build_ui(self):
        for child in self.winfo_children():
            if child is getattr(self, "_splash", None):
                continue
            child.destroy()
        self.vars = {}
        self.detail = {}
        self._build_style()
        self._build_menubar()
        self._build_header()
        self._build_toolbar()
        self._build_body()
        self._build_statusbar()
        self.path_var.set(str(self.store.path))
        self._bind_shortcuts()

    def toggle_theme(self):
        self.theme = "dark" if self.theme == "light" else "light"
        COL.clear()
        COL.update(PALETTES[self.theme])
        self.style.theme_use(THEME_DARK if self.theme == "dark" else THEME_LIGHT)
        self.configure(bg=COL["bg"])
        cur_id = None
        if self.current_index is not None and \
                0 <= self.current_index < len(self.records):
            cur_id = self.records[self.current_index].get("id")
        self.build_ui()
        self.refresh_tree()
        self.update_dashboard()
        self.update_alert_badge()
        if cur_id is not None:
            self._select_by_id(cur_id)
        self.config_data["theme"] = self.theme
        save_config(self.config_data)
        self.set_status(f"Thème : {self.theme}")


    def _require_login(self) -> bool:
        cfg = self.config_data
        self.role = "admin"
        if not cfg.get("pw_hash"):
            return True
        self.withdraw()
        comp_hash = cfg.get("pw_hash_c")
        for _ in range(3):
            pw = self._ask_password("🔐 Connexion", "Mot de passe :")
            if pw is None:
                return False
            if verify_pw(pw, cfg):
                self.role = "admin"
                self.deiconify()
                return True
            if comp_hash and verify_hash(pw, cfg.get("pw_salt_c", ""), comp_hash,
                                         cfg.get("pw_algo_c", "sha256")):
                self.role = "comptable"
                self.deiconify()
                return True
            messagebox.showerror("خطأ", "كلمة السر غالطة.")
        messagebox.showerror("ممنوع", "3 محاولات فاشلة. البرنامج غادي يتسد.")
        return False

    def _require_admin(self) -> bool:
        if getattr(self, "role", "admin") == "admin":
            return True
        messagebox.showwarning(
            "ممنوع", "هاد العملية مخصّصة للمدير (administrateur) فقط.")
        return False

    def _ask_password(self, title, prompt, confirm=False):
        dlg = tk.Toplevel(self)
        dlg.title(title)
        dlg.configure(bg=COL["surface"])
        dlg.resizable(False, False)
        dlg.grab_set()
        result = {"value": None}

        tk.Label(dlg, text=prompt, bg=COL["surface"], fg=COL["text"],
                 font=(FONT, 11)).pack(padx=24, pady=(20, 6), anchor=tk.W)
        e1 = tk.Entry(dlg, show="•", font=(FONT, 11), width=26)
        e1.pack(padx=24, pady=4)
        e1.focus_set()
        e2 = None
        if confirm:
            tk.Label(dlg, text="Confirmer :", bg=COL["surface"], fg=COL["text"],
                     font=(FONT, 11)).pack(padx=24, pady=(6, 4), anchor=tk.W)
            e2 = tk.Entry(dlg, show="•", font=(FONT, 11), width=26)
            e2.pack(padx=24, pady=4)

        def ok(_e=None):
            if confirm and e1.get() != e2.get():
                messagebox.showwarning("خطأ", "كلمتا السر ماتطابقوش.", parent=dlg)
                return
            result["value"] = e1.get()
            dlg.destroy()

        def cancel(_e=None):
            result["value"] = None
            dlg.destroy()

        btns = tk.Frame(dlg, bg=COL["surface"])
        btns.pack(pady=(10, 18))
        tb.Button(btns, text="OK", bootstyle="success",
                   command=ok).pack(side=tk.LEFT, padx=6)
        tb.Button(btns, text="Annuler", bootstyle="secondary-outline",
                   command=cancel).pack(side=tk.LEFT, padx=6)
        dlg.bind("<Return>", ok)
        dlg.bind("<Escape>", cancel)

        dlg.update_idletasks()
        x = self.winfo_rootx() + 120
        y = self.winfo_rooty() + 120
        dlg.geometry(f"+{x}+{y}")
        self.wait_window(dlg)
        return result["value"]

    def manage_password(self):
        if not self._require_admin():
            return
        cfg = self.config_data
        if cfg.get("pw_hash"):
            cur = self._ask_password("🔐 Mot de passe actuel", "Mot de passe actuel :")
            if cur is None:
                return
            if not verify_pw(cur, cfg):
                messagebox.showerror("خطأ", "كلمة السر الحالية غالطة.")
                return
        new = self._ask_password(
            "🔐 Nouveau mot de passe administrateur",
            "Nouveau mot de passe administrateur\n"
            "(خليه فارغ باش تشيل الحماية) :", confirm=True)
        if new is None:
            return
        if new == "":
            for k in ("pw_hash", "pw_salt", "pw_algo",
                      "pw_hash_c", "pw_salt_c", "pw_algo_c"):
                cfg.pop(k, None)
            save_config(cfg)
            messagebox.showinfo("تم", "تشالات الحماية بكلمة السر.")
            self.set_status("Mot de passe désactivé.")
            return
        cfg.update(make_password(new))
        save_config(cfg)
        messagebox.showinfo("تم", "تسجلات كلمة سر المدير ✓")
        self.set_status("Mot de passe admin activé 🔐")

    def manage_password_comptable(self):
        if not self._require_admin():
            return
        cfg = self.config_data
        if not cfg.get("pw_hash"):
            messagebox.showinfo(
                "أولاً", "خاصك تفعّل كلمة سر المدير الأول باش تستعمل الأدوار.")
            return
        new = self._ask_password(
            "👤 Mot de passe comptable",
            "Mot de passe du comptable (accès limité)\n"
            "(خليه فارغ باش تمسحو) :", confirm=True)
        if new is None:
            return
        if new == "":
            for k in ("pw_hash_c", "pw_salt_c", "pw_algo_c"):
                cfg.pop(k, None)
            save_config(cfg)
            messagebox.showinfo("تم", "تمسح حساب المحاسب.")
            return
        p = make_password(new)
        cfg["pw_salt_c"] = p["pw_salt"]
        cfg["pw_hash_c"] = p["pw_hash"]
        cfg["pw_algo_c"] = p["pw_algo"]
        save_config(cfg)
        messagebox.showinfo(
            "تم", "تسجل حساب المحاسب ✓\nكيقدر يشوف ويصدّر، ولكن ماكيقدرش يمسح "
                  "ولا يبدّل الإعدادات.")


    def _build_style(self):
        # ttkbootstrap gère le thème ; on ajuste juste la hauteur des lignes.
        self.style.configure("Treeview", rowheight=30, font=(FONT, 10))
        self.style.configure("Treeview.Heading", font=(FONT, 10, "bold"))


    def _build_header(self):
        bar = tk.Frame(self, bg=COL["brand"], height=64)
        bar.pack(side=tk.TOP, fill=tk.X)
        bar.pack_propagate(False)

        left = tk.Frame(bar, bg=COL["brand"])
        left.pack(side=tk.LEFT, padx=18)
        tk.Label(left, text="👥  Gestion des Employés & Paie", bg=COL["brand"],
                 fg="white", font=(FONT, 16, "bold")).pack(anchor=tk.W, pady=(10, 0))
        tk.Label(left, text="Fiches du personnel & bulletins de paie reliés à Excel",
                 bg=COL["brand"], fg="#a7d7bd", font=(FONT, 9)).pack(anchor=tk.W)

        right = tk.Frame(bar, bg=COL["brand"])
        right.pack(side=tk.RIGHT, padx=18)
        tk.Label(right, text=COMPANY_NAME, bg=COL["brand"], fg="white",
                 font=(FONT, 14, "bold")).pack(anchor=tk.E, pady=(10, 0))
        role = getattr(self, "role", "admin")
        role_txt = "👑 Administrateur" if role == "admin" else "👤 Comptable"
        tk.Label(right, text=role_txt, bg=COL["brand"], fg="#a7d7bd",
                 font=(FONT, 9)).pack(anchor=tk.E)

    def _build_menubar(self):
        menubar = tk.Menu(self)

        m_file = tk.Menu(menubar, tearoff=0)
        m_file.add_command(label="📂  Ouvrir un fichier Excel…",
                           command=self.choose_file)
        m_file.add_command(label="🔄  Rafraîchir", command=self.reload)
        m_file.add_command(label="📁  Ouvrir le dossier", command=self.open_folder)
        m_file.add_separator()
        m_file.add_command(label="📥  Importer des employés (CSV)…",
                           command=self.import_csv)
        m_file.add_command(label="📤  Exporter les employés (CSV)…",
                           command=self.export_csv)
        m_file.add_separator()
        m_file.add_command(label="Quitter", command=self.destroy)
        menubar.add_cascade(label="Fichier", menu=m_file)

        m_paie = tk.Menu(menubar, tearoff=0)
        m_paie.add_command(label="📅  Pointage…", command=self.open_pointage)
        m_paie.add_command(label="🗓  Calendrier annuel…",
                           command=self.open_year_pointage)
        m_paie.add_command(label="🕐  Heures (entrée/sortie)…",
                           command=self.open_hours)
        m_paie.add_command(label="💳  Avances…", command=self.open_advances)
        m_paie.add_command(label="🏖  Congés…", command=self.open_leaves)
        m_paie.add_command(label="📎  Documents…", command=self.open_documents)
        m_paie.add_command(label="📦  Archiver le mois", command=self.archive_month)
        m_paie.add_separator()
        m_paie.add_command(label="🧾  Bulletin PDF (employé)",
                           command=self.export_pdf)
        m_paie.add_command(label="📚  Tous les bulletins PDF",
                           command=self.export_all_pdf)
        m_paie.add_separator()
        m_paie.add_command(label="📋  Attestation de travail",
                           command=lambda: self.generate_attestation("travail"))
        m_paie.add_command(label="📋  Attestation de salaire",
                           command=lambda: self.generate_attestation("salaire"))
        m_paie.add_command(label="📝  Contrat de travail",
                           command=self.generate_contract)
        menubar.add_cascade(label="Paie", menu=m_paie)

        m_exp = tk.Menu(menubar, tearoff=0)
        m_exp.add_command(label="📊  État de paie (HTML)",
                          command=self.export_etat_paie)
        m_exp.add_command(label="📑  État de paie (Excel)",
                          command=self.export_etat_excel)
        m_exp.add_command(label="🧾  Déclaration CNSS (Excel)",
                          command=self.export_cnss)
        menubar.add_cascade(label="Exporter", menu=m_exp)

        m_view = tk.Menu(menubar, tearoff=0)
        m_view.add_command(label="📈  Graphique des salaires",
                           command=self.open_graph)
        m_view.add_command(label="📉  Évolution mensuelle",
                           command=self.open_evolution)
        m_view.add_command(label="🧮  Simulateur de salaire",
                           command=self.open_simulateur)
        m_view.add_command(label="🗂  Registre des documents",
                           command=self.open_registry)
        m_view.add_command(label="🗄  Employés archivés",
                           command=self.open_archived)
        m_view.add_command(label="📝  Journal des modifications",
                           command=self.open_audit)
        m_view.add_command(label="🔔  Alertes RH", command=self.open_alerts)
        m_view.add_separator()
        theme_lbl = "☀️  Thème clair" if self.theme == "dark" else "🌙  Thème sombre"
        m_view.add_command(label=theme_lbl, command=self.toggle_theme)
        menubar.add_cascade(label="Affichage", menu=m_view)

        m_sec = tk.Menu(menubar, tearoff=0)
        m_sec.add_command(label="🔐  Mot de passe administrateur…",
                          command=self.manage_password)
        m_sec.add_command(label="👤  Mot de passe comptable…",
                          command=self.manage_password_comptable)
        menubar.add_cascade(label="Sécurité", menu=m_sec)

        menubar.add_command(label="⚙️ Paramètres", command=self.open_settings)

        m_help = tk.Menu(menubar, tearoff=0)
        m_help.add_command(label="À propos", command=self._about)
        menubar.add_cascade(label="?", menu=m_help)

        self.config(menu=menubar)

    def _about(self):
        messagebox.showinfo(
            "À propos",
            f"{APP_TITLE}\n\nGestion des employés & paie (Maroc) reliée à Excel.\n"
            f"CNSS / AMO / IR, pointage, congés, bulletins PDF, archive.\n\n"
            f"Société : {COMPANY_NAME}")

    def _rebuild_all(self):
        cur_id = None
        if self.current_index is not None and \
                0 <= self.current_index < len(self.records):
            cur_id = self.records[self.current_index].get("id")
        self.build_ui()
        self.refresh_tree()
        self.update_dashboard()
        self.update_alert_badge()
        if cur_id is not None:
            self._select_by_id(cur_id)


    def open_settings(self):
        if not self._require_admin():
            return
        cur = {
            "company_name": COMPANY_NAME, "logo_path": LOGO_PATH,
            "jours_ouvrables": JOURS_OUVRABLES, "leave_per_month": LEAVE_PER_MONTH,
            "plafond_cnss": PLAFOND_CNSS, "taux_cnss": TAUX_CNSS,
            "taux_amo": TAUX_AMO, "seuil_fp": SEUIL_FP,
            "plafond_fp": PLAFOND_FP_MENSUEL, "deduction_charge": DEDUCTION_CHARGE,
        }
        rows = [
            ("Nom de la société", "company_name", "text"),
            ("Jours ouvrables / mois", "jours_ouvrables", "int"),
            ("Congés acquis / mois (jours)", "leave_per_month", "float"),
            ("Plafond CNSS (DH)", "plafond_cnss", "float"),
            ("Taux CNSS (%)", "taux_cnss", "pct"),
            ("Taux AMO (%)", "taux_amo", "pct"),
            ("Seuil frais pro. (DH)", "seuil_fp", "float"),
            ("Plafond frais pro. / mois (DH)", "plafond_fp", "float"),
            ("Déduction / personne à charge (DH)", "deduction_charge", "float"),
        ]

        dlg = tk.Toplevel(self)
        dlg.title("⚙️ Paramètres")
        dlg.configure(bg=COL["surface"])
        dlg.resizable(False, False)
        dlg.grab_set()

        tk.Label(dlg, text="⚙️  Paramètres", bg=COL["surface"], fg=COL["brand"],
                 font=(FONT, 14, "bold")).grid(row=0, column=0, columnspan=2,
                                               sticky=tk.W, padx=16, pady=(14, 8))
        entries = {}
        for i, (label, key, kind) in enumerate(rows, start=1):
            tk.Label(dlg, text=label, bg=COL["surface"], fg=COL["text"],
                     font=(FONT, 10)).grid(row=i, column=0, sticky=tk.W,
                                           padx=(16, 8), pady=4)
            var = tk.StringVar()
            val = cur[key]
            var.set(f"{val * 100:.2f}" if kind == "pct" else str(val))
            ttk.Entry(dlg, textvariable=var, width=24).grid(
                row=i, column=1, sticky=tk.EW, padx=(0, 16), pady=4)
            entries[key] = (var, kind)

        logo_row = len(rows) + 1
        tk.Label(dlg, text="Logo de la société", bg=COL["surface"], fg=COL["text"],
                 font=(FONT, 10)).grid(row=logo_row, column=0, sticky=tk.W,
                                       padx=(16, 8), pady=4)
        logo_var = tk.StringVar(value=LOGO_PATH)
        logo_box = tk.Frame(dlg, bg=COL["surface"])
        logo_box.grid(row=logo_row, column=1, sticky=tk.EW, padx=(0, 16), pady=4)
        tk.Label(logo_box, textvariable=logo_var, bg=COL["surface"],
                 fg=COL["muted"], font=(FONT, 8), width=22, anchor=tk.W).pack(
                     side=tk.LEFT)

        def pick_logo():
            p = filedialog.askopenfilename(
                title="اختر لوگو", parent=dlg,
                filetypes=[("Images", "*.png *.jpg *.jpeg *.gif *.bmp")])
            if p:
                logo_var.set(p)

        tb.Button(logo_box, text="📷", bootstyle="secondary-outline",
                   command=pick_logo).pack(side=tk.LEFT, padx=4)
        tb.Button(logo_box, text="✖", bootstyle="secondary-outline",
                   command=lambda: logo_var.set("")).pack(side=tk.LEFT)

        sign_row = logo_row + 1
        tk.Label(dlg, text="Signature / cachet", bg=COL["surface"], fg=COL["text"],
                 font=(FONT, 10)).grid(row=sign_row, column=0, sticky=tk.W,
                                       padx=(16, 8), pady=4)
        sign_var = tk.StringVar(value=SIGN_PATH)
        sign_box = tk.Frame(dlg, bg=COL["surface"])
        sign_box.grid(row=sign_row, column=1, sticky=tk.EW, padx=(0, 16), pady=4)
        tk.Label(sign_box, textvariable=sign_var, bg=COL["surface"],
                 fg=COL["muted"], font=(FONT, 8), width=22, anchor=tk.W).pack(
                     side=tk.LEFT)

        def pick_sign():
            p = filedialog.askopenfilename(
                title="اختر التوقيع/الختم", parent=dlg,
                filetypes=[("Images", "*.png *.jpg *.jpeg *.gif *.bmp")])
            if p:
                sign_var.set(p)

        tb.Button(sign_box, text="✍️", bootstyle="secondary-outline",
                   command=pick_sign).pack(side=tk.LEFT, padx=4)
        tb.Button(sign_box, text="✖", bootstyle="secondary-outline",
                   command=lambda: sign_var.set("")).pack(side=tk.LEFT)

        def save():
            updates = {}
            for key, (var, kind) in entries.items():
                raw = var.get().strip()
                if kind == "text":
                    updates[key] = raw
                elif kind == "int":
                    try:
                        updates[key] = int(float(raw))
                    except ValueError:
                        messagebox.showwarning("خطأ", f"قيمة غير صحيحة: {key}",
                                               parent=dlg)
                        return
                else:
                    try:
                        num = float(raw.replace(",", "."))
                    except ValueError:
                        messagebox.showwarning("خطأ", f"قيمة غير صحيحة: {key}",
                                               parent=dlg)
                        return
                    updates[key] = num / 100 if kind == "pct" else num
            updates["logo_path"] = logo_var.get().strip()
            updates["signature_path"] = sign_var.get().strip()
            self.config_data.update(updates)
            save_config(self.config_data)
            apply_config_settings(self.config_data)
            dlg.destroy()
            self._rebuild_all()
            self.set_status("الإعدادات تسجلات ✓")
            messagebox.showinfo("تم", "الإعدادات تطبقات على البرنامج كامل ✓")

        tb.Button(dlg, text="🧮  Éditer le barème IR…", bootstyle="secondary-outline",
                   command=lambda: self.open_ir_editor(dlg)).grid(
                       row=logo_row + 2, column=0, columnspan=2,
                       sticky=tk.W, padx=16, pady=(8, 0))

        btns = tk.Frame(dlg, bg=COL["surface"])
        btns.grid(row=logo_row + 3, column=0, columnspan=2, pady=(12, 16))
        tb.Button(btns, text="💾  Enregistrer", bootstyle="success",
                   command=save).pack(side=tk.LEFT, padx=6)
        tb.Button(btns, text="Annuler", bootstyle="secondary-outline",
                   command=dlg.destroy).pack(side=tk.LEFT, padx=6)
        dlg.columnconfigure(1, weight=1)
        dlg.update_idletasks()
        dlg.geometry(f"+{self.winfo_rootx() + 100}+{self.winfo_rooty() + 60}")

    def open_ir_editor(self, parent=None):
        dlg = tk.Toplevel(parent or self)
        dlg.title("🧮 Barème IR mensuel")
        dlg.configure(bg=COL["surface"])
        dlg.grab_set()
        tk.Label(dlg, text="🧮  Barème IR mensuel", bg=COL["surface"],
                 fg=COL["brand"], font=(FONT, 13, "bold")).pack(
                     anchor=tk.W, padx=16, pady=(14, 4))
        tk.Label(dlg, text="Une tranche par ligne :  plafond ; taux%\n"
                 "Dernière tranche : laisser le plafond vide pour ∞ (sans limite).",
                 bg=COL["surface"], fg=COL["muted"], font=(FONT, 9),
                 justify=tk.LEFT).pack(anchor=tk.W, padx=16)

        txt = tk.Text(dlg, width=30, height=10, font=("Consolas", 11))
        txt.pack(padx=16, pady=8, fill=tk.BOTH, expand=True)
        for cap, rate in IR_BRACKETS:
            cap_s = "" if cap == float("inf") else f"{cap:g}"
            txt.insert(tk.END, f"{cap_s} ; {rate * 100:g}\n")

        def save_ir():
            brackets = []
            for ln in txt.get("1.0", tk.END).strip().splitlines():
                ln = ln.strip()
                if not ln:
                    continue
                parts = ln.replace("%", "").split(";")
                if len(parts) != 2:
                    messagebox.showwarning("خطأ", f"سطر غير صحيح: {ln}", parent=dlg)
                    return
                cap_s, rate_s = parts[0].strip(), parts[1].strip()
                try:
                    cap = None if cap_s == "" else float(cap_s)
                    rate = float(rate_s.replace(",", ".")) / 100
                except ValueError:
                    messagebox.showwarning("خطأ", f"قيمة غير صحيحة: {ln}", parent=dlg)
                    return
                brackets.append([cap, rate])
            if not brackets:
                messagebox.showwarning("فارغ", "خاصك على الأقل تَرانش وحدة.", parent=dlg)
                return
            self.config_data["ir_brackets"] = brackets
            save_config(self.config_data)
            apply_config_settings(self.config_data)
            dlg.destroy()
            messagebox.showinfo("تم", "جدول IR تسجل وتطبّق ✓")
            self._rebuild_all()

        bb = tk.Frame(dlg, bg=COL["surface"])
        bb.pack(pady=(0, 14))
        tb.Button(bb, text="💾  Enregistrer", bootstyle="success",
                   command=save_ir).pack(side=tk.LEFT, padx=6)
        tb.Button(bb, text="Annuler", bootstyle="secondary-outline",
                   command=dlg.destroy).pack(side=tk.LEFT, padx=6)
        dlg.geometry(f"+{self.winfo_rootx() + 140}+{self.winfo_rooty() + 90}")

    def _build_toolbar(self):
        bar = tk.Frame(self, bg=COL["surface"], height=52,
                       highlightbackground=COL["border"], highlightthickness=1)
        bar.pack(side=tk.TOP, fill=tk.X)
        bar.pack_propagate(False)

        quick = [("📂  Fichier", self.choose_file, "Ouvrir un fichier Excel"),
                 ("🔄  Rafraîchir", self.reload, "Recharger les données"),
                 ("📅  Pointage", self.open_pointage, "Pointage mensuel"),
                 ("🧾  PDF", self.export_pdf, "Bulletin de paie PDF (Ctrl+P)"),
                 ("📚  Tous PDF", self.export_all_pdf, "Tous les bulletins PDF"),
                 ("📉  Évolution", self.open_evolution, "Évolution de la masse salariale")]
        first = True
        for label, cmd, tip in quick:
            b = tb.Button(bar, text=label, bootstyle="light", command=cmd)
            b.pack(side=tk.LEFT, padx=((12, 4) if first else 4), pady=8)
            self._tip(b, tip)
            first = False

        self.path_var = tk.StringVar()
        theme_icon = "☀️" if self.theme == "dark" else "🌙"
        bt = tb.Button(bar, text=theme_icon, bootstyle="light",
                       command=self.toggle_theme)
        bt.pack(side=tk.RIGHT, padx=4, pady=8)
        self._tip(bt, "Thème clair / sombre")
        bp = tb.Button(bar, text="🔐", bootstyle="light",
                       command=self.manage_password)
        bp.pack(side=tk.RIGHT, padx=4, pady=8)
        self._tip(bp, "Mot de passe administrateur")
        self.alert_btn = tb.Button(bar, text="🔔  Alertes", bootstyle="warning",
                                    command=self.open_alerts)
        self.alert_btn.pack(side=tk.RIGHT, padx=4, pady=8)
        self._tip(self.alert_btn, "Alertes RH (anniversaires, fins de contrat)")

    def _tip(self, widget, text):
        if HAS_TB_EXTRAS:
            try:
                ToolTip(widget, text=text, bootstyle="secondary-inverse")
            except Exception:  # noqa: BLE001
                pass


    def _build_dashboard(self, parent):
        row = tk.Frame(parent, bg=COL["bg"])
        row.pack(fill=tk.X, pady=(0, 12))

        self.stat_effectif = tk.StringVar(value="0")
        self.stat_brut = tk.StringVar(value="0,00 DH")
        self.stat_masse = tk.StringVar(value="0,00 DH")
        self.stat_moyen = tk.StringVar(value="0,00 DH")
        self.stat_ir = tk.StringVar(value="0,00 DH")

        chips = [
            ("👥", "Effectif", self.stat_effectif, COL["info"]),
            ("💰", "Masse brute", self.stat_brut, COL["brand2"]),
            ("💵", "Masse nette", self.stat_masse, COL["accent"]),
            ("📊", "Net moyen", self.stat_moyen, COL["brand2"]),
            ("🧾", "Total IR", self.stat_ir, COL["danger"]),
        ]
        for icon, title, var, color in chips:
            chip = tk.Frame(row, bg=COL["surface"], highlightbackground=COL["border"],
                            highlightthickness=1)
            chip.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
            inner = tk.Frame(chip, bg=COL["surface"])
            inner.pack(fill=tk.X, padx=14, pady=10)
            tk.Label(inner, text=icon, bg=COL["surface"],
                     font=(FONT, 18)).pack(side=tk.LEFT, padx=(0, 10))
            txt = tk.Frame(inner, bg=COL["surface"])
            txt.pack(side=tk.LEFT)
            tk.Label(txt, text=title, bg=COL["surface"], fg=COL["muted"],
                     font=(FONT, 9)).pack(anchor=tk.W)
            tk.Label(txt, textvariable=var, bg=COL["surface"], fg=color,
                     font=(FONT, 14, "bold")).pack(anchor=tk.W)

    def update_dashboard(self):
        payrolls = [compute_payroll(r) for r in self.records]
        n = len(payrolls)
        total_net = sum(p["net"] for p in payrolls)
        total_brut = sum(p["brut"] for p in payrolls)
        total_ir = sum(p["ir"] for p in payrolls)
        avg = total_net / n if n else 0
        self.stat_effectif.set(str(n))
        self.stat_brut.set(fmt_money(total_brut))
        self.stat_masse.set(fmt_money(total_net))
        self.stat_moyen.set(fmt_money(avg))
        self.stat_ir.set(fmt_money(total_ir))


    def _build_body(self):
        body = tk.Frame(self, bg=COL["bg"])
        body.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)

        self._build_left(body)

        right = tk.Frame(body, bg=COL["bg"])
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._build_dashboard(right)
        self._build_profile(right)
        self._build_scroll_form(right)
        self._build_actions(right)

    def _build_left(self, parent):
        card = tk.Frame(parent, bg=COL["surface"],
                        highlightbackground=COL["border"], highlightthickness=1)
        card.pack(side=tk.LEFT, fill=tk.BOTH, expand=False, padx=(0, 12))
        card.configure(width=360)
        card.pack_propagate(False)

        head = tk.Frame(card, bg=COL["surface"])
        head.pack(fill=tk.X, padx=12, pady=(12, 6))
        self.count_var = tk.StringVar(value="Employés")
        tk.Label(head, textvariable=self.count_var, bg=COL["surface"],
                 fg=COL["text"], font=(FONT, 12, "bold")).pack(side=tk.LEFT)

        sbox = tk.Frame(card, bg="#f1f5f9", highlightbackground=COL["border"],
                        highlightthickness=1)
        sbox.pack(fill=tk.X, padx=12, pady=(0, 6))
        tk.Label(sbox, text="🔎", bg="#f1f5f9").pack(side=tk.LEFT, padx=(8, 2))
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *_: self.refresh_tree())
        self.search_entry = tk.Entry(sbox, textvariable=self.search_var, bd=0,
                                     bg="#f1f5f9", font=(FONT, 10), fg=COL["text"])
        self.search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=6,
                               padx=(0, 8))

        filt = tk.Frame(card, bg=COL["surface"])
        filt.pack(fill=tk.X, padx=12, pady=(0, 8))
        self.filter_poste = tk.StringVar(value="Tous postes")
        self.filter_contrat = tk.StringVar(value="Tous contrats")
        self.poste_combo = ttk.Combobox(
            filt, textvariable=self.filter_poste, state="readonly",
            values=["Tous postes"], font=(FONT, 9))
        self.poste_combo.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4))
        contrat_combo = ttk.Combobox(
            filt, textvariable=self.filter_contrat, state="readonly",
            values=["Tous contrats"] + CONTRAT_OPTIONS, font=(FONT, 9), width=12)
        contrat_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.filter_poste.trace_add("write", lambda *_: self.refresh_tree())
        self.filter_contrat.trace_add("write", lambda *_: self.refresh_tree())

        wrap = tk.Frame(card, bg=COL["surface"])
        wrap.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 12))
        cols = ("id", "nom", "poste", "salaire_net")
        self.tree = ttk.Treeview(wrap, columns=cols, show="headings",
                                 selectmode="browse")
        for col, label, width, anchor in (
            ("id", "ID", 42, tk.CENTER),
            ("nom", "Nom", 150, tk.W),
            ("poste", "Poste", 95, tk.W),
            ("salaire_net", "Net", 90, tk.E),
        ):
            self.tree.heading(col, text=label)
            self.tree.column(col, width=width, anchor=anchor)
        self.tree.tag_configure("odd", background=COL["row_alt"])
        self.tree.tag_configure("even", background=COL["surface"])
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb = ttk.Scrollbar(wrap, orient="vertical", command=self.tree.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.bind("<<TreeviewSelect>>", self.on_select)

    def _build_profile(self, parent):
        card = tk.Frame(parent, bg=COL["surface"],
                        highlightbackground=COL["border"], highlightthickness=1)
        card.pack(fill=tk.X)
        inner = tk.Frame(card, bg=COL["surface"])
        inner.pack(fill=tk.X, padx=16, pady=14)

        self.vars["photo"] = tk.StringVar()
        self._photo_img = None

        photo_box = tk.Frame(inner, bg=COL["surface"])
        photo_box.pack(side=tk.LEFT)
        self.avatar = tk.Canvas(photo_box, width=64, height=64, bg=COL["surface"],
                                highlightthickness=0)
        self.avatar.pack()
        btns = tk.Frame(photo_box, bg=COL["surface"])
        btns.pack(pady=(4, 0))
        tk.Button(btns, text="📷", bd=0, bg=COL["surface"], cursor="hand2",
                  activebackground=COL["bg"], font=(FONT, 9),
                  command=self.choose_photo).pack(side=tk.LEFT)
        tk.Button(btns, text="✖", bd=0, bg=COL["surface"], cursor="hand2",
                  activebackground=COL["bg"], fg=COL["muted"], font=(FONT, 9),
                  command=self.clear_photo).pack(side=tk.LEFT, padx=(6, 0))

        info = tk.Frame(inner, bg=COL["surface"])
        info.pack(side=tk.LEFT, padx=14)
        self.name_var = tk.StringVar(value="—")
        self.poste_var = tk.StringVar(value="Aucun employé sélectionné")
        tk.Label(info, textvariable=self.name_var, bg=COL["surface"],
                 fg=COL["text"], font=(FONT, 16, "bold")).pack(anchor=tk.W)
        tk.Label(info, textvariable=self.poste_var, bg=COL["surface"],
                 fg=COL["muted"], font=(FONT, 10)).pack(anchor=tk.W)

        badge = tk.Frame(inner, bg=COL["net_bg"], highlightbackground=COL["accent"],
                         highlightthickness=1)
        badge.pack(side=tk.RIGHT)
        tk.Label(badge, text="SALAIRE NET", bg=COL["net_bg"], fg=COL["brand2"],
                 font=(FONT, 8, "bold")).pack(anchor=tk.E, padx=14, pady=(8, 0))
        self.net_var = tk.StringVar(value="—")
        tk.Label(badge, textvariable=self.net_var, bg=COL["net_bg"],
                 fg=COL["brand2"], font=(FONT, 17, "bold")).pack(
                     anchor=tk.E, padx=14, pady=(0, 8))
        self.render_avatar("")

    def render_avatar(self, name: str):
        self.avatar.delete("all")
        path = self.vars["photo"].get().strip()
        if path and HAS_PIL and os.path.exists(path):
            try:
                img = Image.open(path).convert("RGBA").resize(
                    (64, 64), Image.LANCZOS)
                mask = Image.new("L", (64, 64), 0)
                ImageDraw.Draw(mask).ellipse((0, 0, 63, 63), fill=255)
                img.putalpha(mask)
                self._photo_img = ImageTk.PhotoImage(img)
                self.avatar.create_image(32, 32, image=self._photo_img)
                return
            except Exception:
                pass
        color = AVATAR_COLORS[abs(hash(name)) % len(AVATAR_COLORS)] if name \
            else COL["muted"]
        self.avatar.create_oval(2, 2, 62, 62, fill=color, outline=color)
        self.avatar.create_text(32, 32, text=initials(name) if name else "—",
                                fill="white", font=(FONT, 20, "bold"))

    def choose_photo(self):
        if not HAS_PIL:
            messagebox.showinfo(
                "Pillow ناقصة",
                "باش تستعمل الصور ثبت Pillow:\n\n    pip install Pillow")
            return
        path = filedialog.askopenfilename(
            title="اختر صورة الخدام",
            filetypes=[("Images", "*.png *.jpg *.jpeg *.gif *.bmp"),
                       ("Tous", "*.*")])
        if path:
            self.vars["photo"].set(path)
            self.render_avatar(self.vars["nom"].get())

    def clear_photo(self):
        self.vars["photo"].set("")
        self.render_avatar(self.vars["nom"].get())

    def _build_scroll_form(self, parent):
        card = tk.Frame(parent, bg=COL["surface"],
                        highlightbackground=COL["border"], highlightthickness=1)
        card.pack(fill=tk.BOTH, expand=True, pady=(12, 0))

        canvas = tk.Canvas(card, bg=COL["surface"], highlightthickness=0)
        scroll = ttk.Scrollbar(card, orient="vertical", command=canvas.yview)
        form = tk.Frame(canvas, bg=COL["surface"])
        form.bind("<Configure>",
                  lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        win = canvas.create_window((0, 0), window=form, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfigure(win, width=e.width))
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.bind_all("<MouseWheel>",
                        lambda e: canvas.yview_scroll(int(-e.delta / 120), "units"))

        self._build_form_fields(form)
        self._build_detail(form)

    def _card(self, parent, title, icon):
        card = tk.Frame(parent, bg=COL["surface"],
                        highlightbackground=COL["border"], highlightthickness=1)
        card.pack(fill=tk.X, padx=14, pady=(14, 0))
        htitle = tk.Frame(card, bg=COL["surface"])
        htitle.pack(fill=tk.X, padx=14, pady=(10, 2))
        tk.Label(htitle, text=f"{icon}  {title}", bg=COL["surface"],
                 fg=COL["brand2"], font=(FONT, 11, "bold")).pack(anchor=tk.W)
        tk.Frame(card, bg=COL["border"], height=1).pack(fill=tk.X, padx=14)
        body = tk.Frame(card, bg=COL["surface"])
        body.pack(fill=tk.X, padx=14, pady=10)
        return body

    def _build_form_fields(self, parent):
        for gkey, (gtitle, icon) in GROUPS.items():
            body = self._card(parent, gtitle, icon)
            body.columnconfigure(1, weight=1)
            body.columnconfigure(3, weight=1)
            cells = [(k, l, kind) for k, l, g, kind in FIELDS if g == gkey]
            for i, (key, label, kind) in enumerate(cells):
                r, c = divmod(i, 2)
                col = c * 2
                tk.Label(body, text=label, bg=COL["surface"], fg=COL["muted"],
                         font=(FONT, 9)).grid(row=r * 2, column=col, columnspan=2,
                                              sticky=tk.W, padx=(4, 8), pady=(4, 0))
                var = tk.StringVar()
                self.vars[key] = var
                if kind == "combo":
                    w = ttk.Combobox(body, textvariable=var,
                                     values=CONTRAT_OPTIONS, state="readonly")
                else:
                    w = ttk.Entry(body, textvariable=var)
                    if kind == "readonly":
                        w.configure(state="readonly")
                w.grid(row=r * 2 + 1, column=col, columnspan=2, sticky=tk.EW,
                       padx=(4, 8), pady=(0, 6))

        for key in FIELD_BY_KEY:
            self.vars.setdefault(key, tk.StringVar())

        for key in ("salaire_base", "primes", "retenues", "jours_absence",
                    "personnes_charge", "date_embauche", "nom", "poste"):
            self.vars[key].trace_add("write", lambda *_: self.update_calc())

    def _build_detail(self, parent):
        body = self._card(parent, "Détail de la paie (calcul automatique)", "🧮")
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=1)

        gains = tk.Frame(body, bg=COL["surface"])
        gains.grid(row=0, column=0, sticky=tk.NSEW, padx=(0, 8))
        tk.Label(gains, text="GAINS", bg=COL["surface"], fg=COL["brand2"],
                 font=(FONT, 9, "bold")).pack(anchor=tk.W, pady=(0, 4))
        self._detail_line(gains, "base", "Salaire de base")
        self._detail_line(gains, "primes", "Primes")
        self._detail_line(gains, "brut", "Salaire brut", strong=True)

        rets = tk.Frame(body, bg=COL["surface"])
        rets.grid(row=0, column=1, sticky=tk.NSEW, padx=(8, 0))
        tk.Label(rets, text="RETENUES", bg=COL["surface"], fg=COL["ded_fg"],
                 font=(FONT, 9, "bold")).pack(anchor=tk.W, pady=(0, 4))
        self._detail_line(rets, "cnss", "CNSS (4.48%)", ded=True)
        self._detail_line(rets, "amo", "AMO (2.26%)", ded=True)
        self._detail_line(rets, "ir", "IR (impôt)", ded=True)
        self._detail_line(rets, "ded_absence", "Retenue absence", ded=True)
        self._detail_line(rets, "autres", "Autres retenues", ded=True)
        self._detail_line(rets, "avance", "Retenue avance", ded=True)
        self._detail_line(rets, "total_retenues", "Total retenues",
                          strong=True, ded=True)

        netrow = tk.Frame(body, bg=COL["net_bg"])
        netrow.grid(row=1, column=0, columnspan=2, sticky=tk.EW, pady=(10, 2))
        tk.Label(netrow, text="💵  SALAIRE NET À PAYER", bg=COL["net_bg"],
                 fg=COL["brand2"], font=(FONT, 10, "bold")).pack(
                     side=tk.LEFT, padx=12, pady=8)
        self.detail["net"] = tk.StringVar(value="—")
        tk.Label(netrow, textvariable=self.detail["net"], bg=COL["net_bg"],
                 fg=COL["brand2"], font=(FONT, 15, "bold")).pack(
                     side=tk.RIGHT, padx=12)

        meta = tk.Frame(body, bg=COL["surface"])
        meta.grid(row=2, column=0, columnspan=2, sticky=tk.EW, pady=(8, 0))
        self.detail["id"] = tk.StringVar()
        self.detail["anciennete"] = tk.StringVar()
        tk.Label(meta, text="ID :", bg=COL["surface"], fg=COL["muted"],
                 font=(FONT, 9)).pack(side=tk.LEFT)
        tk.Label(meta, textvariable=self.detail["id"], bg=COL["surface"],
                 fg=COL["text"], font=(FONT, 9, "bold")).pack(side=tk.LEFT, padx=(4, 18))
        tk.Label(meta, text="Ancienneté :", bg=COL["surface"], fg=COL["muted"],
                 font=(FONT, 9)).pack(side=tk.LEFT)
        tk.Label(meta, textvariable=self.detail["anciennete"], bg=COL["surface"],
                 fg=COL["text"], font=(FONT, 9, "bold")).pack(side=tk.LEFT, padx=4)

        cong = tk.Frame(body, bg=COL["net_bg"])
        cong.grid(row=3, column=0, columnspan=2, sticky=tk.EW, pady=(8, 0))
        tk.Label(cong, text="🏖  Congés", bg=COL["net_bg"], fg=COL["brand2"],
                 font=(FONT, 9, "bold")).pack(side=tk.LEFT, padx=12, pady=6)
        self.detail["conges"] = tk.StringVar(value="—")
        tk.Label(cong, textvariable=self.detail["conges"], bg=COL["net_bg"],
                 fg=COL["text"], font=(FONT, 9, "bold")).pack(side=tk.RIGHT, padx=12)

        tk.Label(body, text="⚠ Taux CNSS/AMO/IR paramétrables en haut du fichier — "
                 "à valider avec votre comptable.", bg=COL["surface"],
                 fg=COL["muted"], font=(FONT, 8)).grid(
                     row=4, column=0, columnspan=2, sticky=tk.W, pady=(8, 0))

    def _detail_line(self, parent, key, label, strong=False, ded=False):
        font = (FONT, 10, "bold") if strong else (FONT, 9)
        fg = COL["text"] if not ded else COL["ded_fg"]
        if strong:
            tk.Frame(parent, bg=COL["border"], height=1).pack(
                fill=tk.X, pady=(2, 2))
        line = tk.Frame(parent, bg=COL["surface"])
        line.pack(fill=tk.X, pady=1)
        tk.Label(line, text=label, bg=COL["surface"], fg=COL["muted"],
                 font=font).pack(side=tk.LEFT)
        var = tk.StringVar(value="0,00 DH")
        self.detail[key] = var
        tk.Label(line, textvariable=var, bg=COL["surface"],
                 fg=fg if not strong else COL["text"], font=font).pack(side=tk.RIGHT)

    def _build_actions(self, parent):
        bar = tk.Frame(parent, bg=COL["bg"])
        bar.pack(fill=tk.X, pady=(10, 0))
        tb.Button(bar, text="🆕  Nouveau", bootstyle="info",
                   command=self.new_record).pack(side=tk.LEFT)
        tb.Button(bar, text="💾  Enregistrer", bootstyle="success",
                   command=self.save_record).pack(side=tk.LEFT, padx=8)
        tb.Button(bar, text="🗄  Archiver", bootstyle="warning",
                   command=self.archive_record).pack(side=tk.LEFT)
        tb.Button(bar, text="🖨  Bulletin de paie", bootstyle="secondary-outline",
                   command=self.export_fiche).pack(side=tk.RIGHT)
        tb.Button(bar, text="📎  Documents", bootstyle="secondary-outline",
                   command=self.open_documents).pack(side=tk.RIGHT, padx=8)
        tb.Button(bar, text="💳  Avances", bootstyle="secondary-outline",
                   command=self.open_advances).pack(side=tk.RIGHT)
        tb.Button(bar, text="🏖  Congés", bootstyle="secondary-outline",
                   command=self.open_leaves).pack(side=tk.RIGHT, padx=6)

    def _build_statusbar(self):
        self.status_var = tk.StringVar()
        bar = tk.Frame(self, bg=COL["brand"], height=26)
        bar.pack(side=tk.BOTTOM, fill=tk.X)
        bar.pack_propagate(False)
        tk.Label(bar, textvariable=self.status_var, bg=COL["brand"],
                 fg="#cfeede", font=(FONT, 9), anchor=tk.W).pack(side=tk.LEFT, padx=12)


    def choose_file(self):
        path = filedialog.askopenfilename(
            title="اختر ملف Excel",
            filetypes=[("Excel", "*.xlsx"), ("Tous", "*.*")],
            initialdir=str(DEFAULT_FILE.parent))
        if not path:
            path = filedialog.asksaveasfilename(
                title="صاوب ملف Excel جديد", defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialdir=str(DEFAULT_FILE.parent), initialfile="employes.xlsx")
            if not path:
                return
        self.store = ExcelStore(Path(path))
        self.pointage = PointageStore(Path(path))
        self.history = HistoryStore(Path(path))
        self.advances = AdvancesStore(Path(path))
        self.audit = AuditStore(Path(path))
        self.leaves = LeaveStore(Path(path))
        self.hours = HoursStore(Path(path))
        self.reload()

    def reload(self):
        try:
            self.records = self.store.load()
        except Exception as exc:
            messagebox.showerror("خطأ فالقراءة", str(exc))
            self.records = []
        self.path_var.set(str(self.store.path))
        self.current_index = None
        self._refresh_poste_options()
        self.clear_form()
        self.refresh_tree()
        self.update_dashboard()
        self.update_alert_badge()
        self.set_status(f"{len(self.records)} خدام محمّلين.")

    def _refresh_poste_options(self):
        if not hasattr(self, "poste_combo"):
            return
        postes = sorted({str(r.get("poste", "")).strip()
                         for r in self.records if str(r.get("poste", "")).strip()})
        self.poste_combo["values"] = ["Tous postes"] + postes
        if self.filter_poste.get() not in self.poste_combo["values"]:
            self.filter_poste.set("Tous postes")

    def refresh_tree(self):
        query = self.search_var.get().strip().lower()
        f_poste = getattr(self, "filter_poste", None)
        f_contrat = getattr(self, "filter_contrat", None)
        f_poste = f_poste.get() if f_poste else "Tous postes"
        f_contrat = f_contrat.get() if f_contrat else "Tous contrats"
        self.tree.delete(*self.tree.get_children())
        shown = 0
        for idx, rec in enumerate(self.records):
            if str(rec.get("archive", "")).strip():
                continue
            haystack = " ".join(str(rec.get(k, "")) for k in
                                ("id", "nom", "cin", "poste", "telephone",
                                 "email")).lower()
            if query and query not in haystack:
                continue
            if f_poste != "Tous postes" and \
                    str(rec.get("poste", "")).strip() != f_poste:
                continue
            if f_contrat != "Tous contrats" and \
                    str(rec.get("type_contrat", "")).strip() != f_contrat:
                continue
            net = compute_payroll(rec)["net"]
            tag = "odd" if shown % 2 else "even"
            self.tree.insert("", tk.END, iid=str(idx), tags=(tag,), values=(
                rec.get("id", ""), rec.get("nom", ""),
                rec.get("poste", ""), fmt_money(net)))
            shown += 1
        self.count_var.set(f"Employés ({shown})")

    def on_select(self, _event=None):
        sel = self.tree.selection()
        if not sel:
            return
        idx = int(sel[0])
        self.current_index = idx
        rec = self.records[idx]
        for key, var in self.vars.items():
            var.set("" if rec.get(key) in (None,) else str(rec.get(key, "")))
        self.update_calc()
        self.set_status(f"الفيش ديال: {rec.get('nom', '')}")

    def clear_form(self):
        for var in self.vars.values():
            var.set("")
        self.update_calc()

    def collect_form(self) -> dict:
        return {key: self.vars[key].get().strip() for key in FIELD_BY_KEY}

    def update_calc(self):
        rec = self.collect_form()
        p = compute_payroll(rec)
        anc = compute_anciennete(rec.get("date_embauche"))

        self.vars["salaire_brut"].set(f"{p['brut']:.2f}")
        self.vars["ret_cnss"].set(f"{p['cnss']:.2f}")
        self.vars["ret_amo"].set(f"{p['amo']:.2f}")
        self.vars["ret_ir"].set(f"{p['ir']:.2f}")
        self.vars["anciennete"].set(anc)
        self.vars["salaire_net"].set(f"{p['net']:.2f}")

        has_data = bool(rec.get("nom") or rec.get("salaire_base"))
        self.name_var.set(rec.get("nom") or "—")
        self.poste_var.set(rec.get("poste") or
                           ("—" if has_data else "Aucun employé sélectionné"))
        self.net_var.set(fmt_money(p["net"]) if has_data else "—")
        self.render_avatar(rec.get("nom") or "")

        for k in ("base", "primes", "brut", "cnss", "amo", "ir",
                  "ded_absence", "autres", "avance", "total_retenues"):
            self.detail[k].set(fmt_money(p[k]))
        self.detail["net"].set(fmt_money(p["net"]) if has_data else "—")
        self.detail["id"].set(rec.get("id") or "—")
        self.detail["anciennete"].set(anc or "—")
        if has_data:
            acq, pris, solde = self._conges(rec)
            self.detail["conges"].set(
                f"Acquis {acq} j   •   Pris {pris} j   •   Solde {solde} j")
        else:
            self.detail["conges"].set("—")

    def next_id(self) -> int:
        ids = [int(to_float(r.get("id"))) for r in self.records
               if str(r.get("id")).strip()]
        return (max(ids) + 1) if ids else 1

    def new_record(self):
        self.current_index = None
        self.tree.selection_remove(self.tree.selection())
        self.clear_form()
        self.vars["id"].set(str(self.next_id()))
        self.vars["date_embauche"].set(dt.date.today().isoformat())
        self.set_status("خدام جديد — عمر المعلومات وسولد 💾 Enregistrer.")

    def save_record(self):
        rec = self.collect_form()
        if not rec.get("nom"):
            messagebox.showwarning("ناقص", "خاصك تكتب على الأقل الاسم الكامل.")
            return

        errors = []
        num_fields = [("salaire_base", "Salaire de base"), ("primes", "Primes"),
                      ("retenues", "Autres retenues"),
                      ("retenue_avance", "Retenue avance"),
                      ("jours_absence", "Jours d'absence"),
                      ("personnes_charge", "Personnes à charge")]
        for key, label in num_fields:
            raw = (rec.get(key) or "").strip()
            if not raw:
                continue
            try:
                val = float(raw.replace(",", ".").replace(" ", ""))
            except ValueError:
                errors.append(f"« {label} » doit être un nombre.")
                continue
            if val < 0:
                errors.append(f"« {label} » ne peut pas être négatif.")
        for key, label in (("date_naissance", "Date de naissance"),
                           ("date_embauche", "Date d'embauche"),
                           ("date_fin_contrat", "Fin de contrat")):
            raw = (rec.get(key) or "").strip()
            if raw and parse_date(raw) is None:
                errors.append(f"« {label} » : date invalide (AAAA-MM-JJ).")
        if errors:
            messagebox.showwarning("Données invalides", "\n".join(errors))
            return

        if not rec.get("id"):
            rec["id"] = str(self.next_id())
        p = compute_payroll(rec)
        rec["salaire_brut"] = f"{p['brut']:.2f}"
        rec["ret_cnss"] = f"{p['cnss']:.2f}"
        rec["ret_amo"] = f"{p['amo']:.2f}"
        rec["ret_ir"] = f"{p['ir']:.2f}"
        rec["anciennete"] = compute_anciennete(rec.get("date_embauche"))
        rec["salaire_net"] = f"{p['net']:.2f}"

        action = "Ajout" if self.current_index is None else "Modification"
        if self.current_index is None:
            self.records.append(rec)
        else:
            self.records[self.current_index] = rec

        try:
            self.store.save_all(self.records)
        except PermissionError:
            messagebox.showerror("الملف محلول",
                                 "ملف Excel محلول. سدّو فـExcel وعاود حاول.")
            return
        except Exception as exc:
            messagebox.showerror("خطأ فالحفظ", str(exc))
            return

        self.audit.log(getattr(self, "role", "admin"), action,
                       f"{rec.get('nom')} (ID {rec.get('id')})")
        self.reload()
        self._select_by_id(rec["id"])
        self.toast(f"Employé enregistré : {rec.get('nom')}")

    def archive_record(self):
        if self.current_index is None:
            messagebox.showinfo("مكاين والو", "اختار شي خدام باش تأرشفو.")
            return
        rec = self.records[self.current_index]
        if not messagebox.askyesno(
                "Archiver", f"Archiver {rec.get('nom')} ?\n"
                "(récupérable depuis « Employés archivés »)"):
            return
        rec["archive"] = "1"
        try:
            self.store.save_all(self.records)
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("خطأ فالحفظ", str(exc))
            return
        self.audit.log(getattr(self, "role", "admin"), "Archivage",
                       f"{rec.get('nom')} (ID {rec.get('id')})")
        self.reload()
        self.toast(f"Employé archivé : {rec.get('nom')}", "warning")

    def _select_by_id(self, emp_id):
        for idx, rec in enumerate(self.records):
            if str(rec.get("id")) == str(emp_id):
                iid = str(idx)
                if self.tree.exists(iid):
                    self.tree.selection_set(iid)
                    self.tree.see(iid)
                    self.on_select()
                break

    def delete_record(self):
        if not self._require_admin():
            return
        if self.current_index is None:
            messagebox.showinfo("مكاين والو", "اختار شي خدام باش تمسحو.")
            return
        rec = self.records[self.current_index]
        if not messagebox.askyesno("تأكيد المسح",
                                   f"واش بصح بغيتي تمسح: {rec.get('nom')} ؟"):
            return
        info = f"{rec.get('nom')} (ID {rec.get('id')})"
        del self.records[self.current_index]
        try:
            self.store.save_all(self.records)
        except Exception as exc:
            messagebox.showerror("خطأ فالحفظ", str(exc))
            return
        self.audit.log(getattr(self, "role", "admin"), "Suppression", info)
        self.reload()
        self.toast("Employé supprimé", "danger")

    def open_archived(self):
        archived = [(i, r) for i, r in enumerate(self.records)
                    if str(r.get("archive", "")).strip()]
        win = tk.Toplevel(self)
        win.title("Employés archivés")
        win.configure(bg=COL["surface"])
        win.geometry("560x440")
        tk.Label(win, text="🗄  Employés archivés", bg=COL["surface"],
                 fg=COL["brand"], font=(FONT, 14, "bold")).pack(
                     anchor=tk.W, padx=16, pady=(14, 6))
        tree = ttk.Treeview(win, columns=("id", "nom", "poste"),
                            show="headings")
        for c, lab, w in (("id", "ID", 60), ("nom", "Nom", 220),
                          ("poste", "Poste", 160)):
            tree.heading(c, text=lab)
            tree.column(c, width=w)
        tree.pack(fill=tk.BOTH, expand=True, padx=16, pady=8)
        idx_map = {}
        for i, r in archived:
            idx_map[str(i)] = i
            tree.insert("", tk.END, iid=str(i), values=(
                r.get("id", ""), r.get("nom", ""), r.get("poste", "")))

        def restore():
            sel = tree.selection()
            if not sel:
                return
            rec = self.records[idx_map[sel[0]]]
            rec["archive"] = ""
            self.store.save_all(self.records)
            self.audit.log(getattr(self, "role", "admin"), "Restauration",
                           f"{rec.get('nom')} (ID {rec.get('id')})")
            tree.delete(sel[0])
            self.reload()
            self.toast(f"Restauré : {rec.get('nom')}", "success")

        def purge():
            if not self._require_admin():
                return
            sel = tree.selection()
            if not sel:
                return
            rec = self.records[idx_map[sel[0]]]
            if not messagebox.askyesno(
                    "Suppression définitive",
                    f"Supprimer définitivement {rec.get('nom')} ?", parent=win):
                return
            info = f"{rec.get('nom')} (ID {rec.get('id')})"
            self.records.pop(idx_map[sel[0]])
            self.store.save_all(self.records)
            self.audit.log(getattr(self, "role", "admin"),
                           "Suppression définitive", info)
            win.destroy()
            self.reload()
            self.toast("Supprimé définitivement", "danger")

        bb = tk.Frame(win, bg=COL["surface"])
        bb.pack(fill=tk.X, padx=16, pady=(0, 14))
        tb.Button(bb, text="↩  Restaurer", bootstyle="success",
                  command=restore).pack(side=tk.LEFT)
        tb.Button(bb, text="🗑  Supprimer définitivement", bootstyle="danger",
                  command=purge).pack(side=tk.LEFT, padx=6)

    def open_audit(self):
        entries = self.audit.entries()
        win = tk.Toplevel(self)
        win.title("Journal des modifications")
        win.configure(bg=COL["surface"])
        win.geometry("680x460")
        tk.Label(win, text="📝  Journal des modifications (Audit)",
                 bg=COL["surface"], fg=COL["brand"], font=(FONT, 14, "bold")).pack(
                     anchor=tk.W, padx=16, pady=(14, 6))
        tree = ttk.Treeview(win, columns=("time", "user", "action", "detail"),
                            show="headings")
        for c, lab, w in (("time", "Date/heure", 150), ("user", "Utilisateur", 90),
                          ("action", "Action", 130), ("detail", "Détail", 260)):
            tree.heading(c, text=lab)
            tree.column(c, width=w, anchor=tk.W)
        tree.pack(fill=tk.BOTH, expand=True, padx=16, pady=8)
        for e in reversed(entries):
            tree.insert("", tk.END, values=(
                e.get("time", "").replace("T", " "), e.get("user", ""),
                e.get("action", ""), e.get("detail", "")))
        tk.Label(win, text=f"{len(entries)} opération(s) — journal en lecture seule",
                 bg=COL["surface"], fg=COL["muted"], font=(FONT, 9)).pack(
                     anchor=tk.W, padx=16, pady=(0, 12))


    def export_fiche(self):
        rec = self.collect_form()
        if not rec.get("nom"):
            messagebox.showwarning("ناقص", "اختار شي خدام أو عمر الفيش الأول.")
            return
        rec["anciennete"] = compute_anciennete(rec.get("date_embauche"))
        out_dir = self.store.path.parent / "fiches"
        out_dir.mkdir(exist_ok=True)
        safe = "".join(c for c in str(rec.get("nom")) if c.isalnum()
                       or c in " _-").strip().replace(" ", "_")
        out_file = out_dir / f"bulletin_{rec.get('id','')}_{safe}.html"
        out_file.write_text(self._bulletin_html(rec), encoding="utf-8")
        webbrowser.open(out_file.as_uri())
        self.set_status(f"البولتان تصدّر: {out_file.name}  (Ctrl+P للطباعة/PDF)")

    def _bulletin_html(self, rec: dict) -> str:
        p = compute_payroll(rec)
        mois = dt.date.today().strftime("%m/%Y")
        _, _, solde = self._conges(rec)
        info = [("Matricule", rec.get("id", "")), ("CIN", rec.get("cin", "")),
                ("Poste", rec.get("poste", "")),
                ("N° CNSS", rec.get("cnss", "")),
                ("Date d'embauche", rec.get("date_embauche", "")),
                ("Ancienneté", rec.get("anciennete", "")),
                ("Type de contrat", rec.get("type_contrat", "")),
                ("Personnes à charge", rec.get("personnes_charge", "") or "0"),
                ("Solde congés", f"{solde} j")]
        info_rows = "".join(
            f"<tr><td class='lbl'>{k}</td><td>{v}</td></tr>" for k, v in info)

        photo_path = str(rec.get("photo", "")).strip()
        photo_html = ""
        if photo_path and os.path.exists(photo_path):
            photo_html = f"<img src='{Path(photo_path).as_uri()}' alt='photo'>"
        logo_html = ""
        if LOGO_PATH and os.path.exists(LOGO_PATH):
            logo_html = f"<img class='logo' src='{Path(LOGO_PATH).as_uri()}'>"

        def line(label, gain="", ret=""):
            return (f"<tr><td>{label}</td>"
                    f"<td class='num'>{gain}</td>"
                    f"<td class='num red'>{ret}</td></tr>")

        lines = (
            line("Salaire de base", fmt_money(p["base"])) +
            line("Primes / indemnités", fmt_money(p["primes"])) +
            line("CNSS (4.48%)", "", fmt_money(p["cnss"])) +
            line("AMO (2.26%)", "", fmt_money(p["amo"])) +
            line("IR (impôt sur le revenu)", "", fmt_money(p["ir"])) +
            line("Retenue absence", "", fmt_money(p["ded_absence"])) +
            line("Autres retenues", "", fmt_money(p["autres"])) +
            line("Retenue avance", "", fmt_money(p["avance"])))

        return f"""<!DOCTYPE html>
<html lang="fr"><head><meta charset="utf-8">
<title>Bulletin — {rec.get('nom','')}</title>
<style>
  *{{box-sizing:border-box;}}
  body{{font-family:"Segoe UI",Arial,sans-serif;color:#1e293b;max-width:780px;
        margin:24px auto;padding:0 16px;}}
  .head{{display:flex;justify-content:space-between;align-items:center;
         background:{COL['brand']};color:#fff;padding:18px 22px;border-radius:10px;}}
  .head h1{{margin:0;font-size:20px;}} .head .co{{font-size:18px;font-weight:bold;}}
  .head .right{{display:flex;align-items:center;gap:10px;}}
  .head .logo{{height:46px;width:auto;background:#fff;border-radius:6px;padding:3px;}}
  .sub{{color:#a7d7bd;font-size:12px;}}
  .emp{{display:flex;align-items:center;gap:14px;margin:18px 0 6px;}}
  .emp img{{width:74px;height:74px;border-radius:50%;object-fit:cover;
            border:3px solid {COL['brand2']};}}
  h2{{margin:0;font-size:20px;}}
  table{{width:100%;border-collapse:collapse;margin-bottom:10px;}}
  td,th{{border:1px solid #e2e8f0;padding:7px 12px;font-size:13px;}}
  td.lbl{{background:#f1f5f9;width:40%;font-weight:600;color:#475569;}}
  th{{background:{COL['brand2']};color:#fff;text-align:left;font-size:12px;}}
  td.num{{text-align:right;}} .red{{color:{COL['ded_fg']};}}
  tr.tot td{{background:#f1f5f9;font-weight:bold;}}
  .net{{margin-top:6px;padding:14px 18px;background:{COL['net_bg']};
        border:1px solid {COL['accent']};border-radius:8px;font-size:19px;
        font-weight:bold;color:{COL['brand2']};text-align:right;}}
  .foot{{margin-top:24px;font-size:11px;color:#94a3b8;text-align:center;}}
  @media print{{.head{{-webkit-print-color-adjust:exact;print-color-adjust:exact;}}}}
</style></head>
<body>
  <div class="head">
    <div><h1>Bulletin de paie</h1><div class="sub">Période : {mois}</div></div>
    <div class="right">{logo_html}<div class="co">{COMPANY_NAME}</div></div>
  </div>
  <div class="emp">{photo_html}<h2>{rec.get('nom','')}</h2></div>
  <table>{info_rows}</table>
  <table>
    <tr><th>Rubrique</th><th style="text-align:right;">Gains</th>
        <th style="text-align:right;">Retenues</th></tr>
    {lines}
    <tr class="tot"><td>Totaux</td>
        <td class="num">{fmt_money(p['brut'])}</td>
        <td class="num red">{fmt_money(p['total_retenues'])}</td></tr>
  </table>
  <div class="net">Net à payer : {fmt_money(p['net'])}</div>
  <div class="foot">Édité le {dt.date.today().isoformat()} — {COMPANY_NAME}<br>
     Document indicatif — cotisations à valider avec le comptable.</div>
</body></html>"""


    def export_etat_paie(self):
        if not self.records:
            messagebox.showinfo("فارغ", "ماكاين حتى خدام باش نصدرو état de paie.")
            return
        out_dir = self.store.path.parent / "fiches"
        out_dir.mkdir(exist_ok=True)
        out_file = out_dir / f"etat_paie_{dt.date.today():%Y_%m}.html"
        out_file.write_text(self._etat_html(), encoding="utf-8")
        webbrowser.open(out_file.as_uri())
        self.set_status(f"État de paie تصدّر: {out_file.name}")

    def _etat_html(self) -> str:
        rows = ""
        tot = {"brut": 0, "cnss": 0, "amo": 0, "ir": 0, "net": 0}
        for rec in self.records:
            p = compute_payroll(rec)
            for k in tot:
                tot[k] += p[k]
            rows += (f"<tr><td>{rec.get('id','')}</td><td>{rec.get('nom','')}</td>"
                     f"<td>{rec.get('poste','')}</td>"
                     f"<td class='num'>{fmt_money(p['brut'])}</td>"
                     f"<td class='num red'>{fmt_money(p['cnss'])}</td>"
                     f"<td class='num red'>{fmt_money(p['amo'])}</td>"
                     f"<td class='num red'>{fmt_money(p['ir'])}</td>"
                     f"<td class='num'><b>{fmt_money(p['net'])}</b></td></tr>")
        mois = dt.date.today().strftime("%m/%Y")
        logo_html = ""
        if LOGO_PATH and os.path.exists(LOGO_PATH):
            logo_html = (f"<img src='{Path(LOGO_PATH).as_uri()}' "
                         f"style='height:42px;background:#fff;border-radius:6px;"
                         f"padding:3px;margin-right:10px;'>")
        return f"""<!DOCTYPE html>
<html lang="fr"><head><meta charset="utf-8">
<title>État de paie {mois}</title>
<style>
  body{{font-family:"Segoe UI",Arial,sans-serif;color:#1e293b;max-width:1000px;
        margin:24px auto;padding:0 16px;}}
  .head{{display:flex;justify-content:space-between;align-items:center;
         background:{COL['brand']};color:#fff;padding:16px 22px;border-radius:10px;}}
  h1{{margin:0;font-size:20px;}}
  table{{width:100%;border-collapse:collapse;margin-top:16px;}}
  td,th{{border:1px solid #e2e8f0;padding:7px 10px;font-size:12px;}}
  th{{background:{COL['brand2']};color:#fff;text-align:left;}}
  td.num{{text-align:right;}} .red{{color:{COL['ded_fg']};}}
  tr.tot td{{background:{COL['net_bg']};font-weight:bold;font-size:13px;}}
  tr:nth-child(even) td{{background:#f8fafc;}}
  .foot{{margin-top:18px;font-size:11px;color:#94a3b8;text-align:center;}}
  @media print{{.head{{-webkit-print-color-adjust:exact;print-color-adjust:exact;}}}}
</style></head>
<body>
  <div class="head"><h1>État de paie — {mois}</h1>
       <div style="display:flex;align-items:center;font-weight:bold;font-size:16px;">
       {logo_html}{COMPANY_NAME}</div></div>
  <table>
    <tr><th>ID</th><th>Nom</th><th>Poste</th>
        <th style="text-align:right;">Brut</th>
        <th style="text-align:right;">CNSS</th>
        <th style="text-align:right;">AMO</th>
        <th style="text-align:right;">IR</th>
        <th style="text-align:right;">Net</th></tr>
    {rows}
    <tr class="tot"><td colspan="3">TOTAUX ({len(self.records)} employés)</td>
        <td class="num">{fmt_money(tot['brut'])}</td>
        <td class="num red">{fmt_money(tot['cnss'])}</td>
        <td class="num red">{fmt_money(tot['amo'])}</td>
        <td class="num red">{fmt_money(tot['ir'])}</td>
        <td class="num">{fmt_money(tot['net'])}</td></tr>
  </table>
  <div class="foot">Édité le {dt.date.today().isoformat()} — {COMPANY_NAME}</div>
</body></html>"""


    def export_etat_excel(self):
        if not self.records:
            messagebox.showinfo("فارغ", "ماكاين حتى خدام باش نصدرو.")
            return
        out_dir = self.store.path.parent / "fiches"
        out_dir.mkdir(exist_ok=True)
        out_file = out_dir / f"etat_paie_{dt.date.today():%Y_%m}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Etat de paie"
        cols = ["ID", "Nom", "Poste", "Brut", "CNSS", "AMO", "IR", "Net"]
        ws.append([f"État de paie — {dt.date.today():%m/%Y} — {COMPANY_NAME}"])
        ws.append(cols)

        head_fill = PatternFill("solid", fgColor="15803D")
        tot_fill = PatternFill("solid", fgColor="EAFAF0")
        white = Font(color="FFFFFF", bold=True)
        bold = Font(bold=True)
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells("A1:H1")
        ws["A1"].font = Font(bold=True, size=13, color="0F3D2E")
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=2, column=c)
            cell.fill = head_fill
            cell.font = white
            cell.border = border

        tot = {"brut": 0, "cnss": 0, "amo": 0, "ir": 0, "net": 0}
        r = 3
        for rec in self.records:
            p = compute_payroll(rec)
            for k in tot:
                tot[k] += p[k]
            values = [rec.get("id", ""), rec.get("nom", ""), rec.get("poste", ""),
                      p["brut"], p["cnss"], p["amo"], p["ir"], p["net"]]
            ws.append(values)
            for c in range(1, len(cols) + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                if c >= 4:
                    cell.number_format = '# ##0.00 "DH"'
            r += 1

        ws.append(["", "", "TOTAUX", tot["brut"], tot["cnss"],
                   tot["amo"], tot["ir"], tot["net"]])
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = tot_fill
            cell.font = bold
            cell.border = border
            if c >= 4:
                cell.number_format = '# ##0.00 "DH"'

        widths = [6, 26, 18, 13, 12, 12, 12, 14]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = w
        ws.freeze_panes = "A3"

        try:
            wb.save(out_file)
        except PermissionError:
            messagebox.showerror("الملف محلول",
                                 "سدّ الملف فـExcel وعاود حاول.")
            return
        try:
            os.startfile(out_file)
        except (AttributeError, OSError):
            webbrowser.open(out_file.as_uri())
        self.set_status(f"État de paie Excel تصدّر: {out_file.name}")

    def export_cnss(self):
        if not self.records:
            messagebox.showinfo("فارغ", "ماكاين حتى خدام.")
            return
        out_dir = self.store.path.parent / "fiches"
        out_dir.mkdir(exist_ok=True)
        out_file = out_dir / f"declaration_cnss_{dt.date.today():%Y_%m}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Declaration CNSS"
        cols = ["N° CNSS", "Nom complet", "CIN", "Jours", "Salaire réel",
                "Salaire plafonné", "CNSS (4.48%)", "AMO (2.26%)"]
        ws.append([f"Déclaration CNSS — {dt.date.today():%m/%Y} — {COMPANY_NAME}"])
        ws.append(cols)

        head_fill = PatternFill("solid", fgColor="15803D")
        tot_fill = PatternFill("solid", fgColor="EAFAF0")
        white = Font(color="FFFFFF", bold=True)
        bold = Font(bold=True)
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells("A1:H1")
        ws["A1"].font = Font(bold=True, size=13, color="0F3D2E")
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=2, column=c)
            cell.fill, cell.font, cell.border = head_fill, white, border

        tot = {"reel": 0.0, "plaf": 0.0, "cnss": 0.0, "amo": 0.0}
        r = 3
        for rec in self.records:
            p = compute_payroll(rec)
            jours = max(0, JOURS_OUVRABLES - int(to_float(rec.get("jours_absence"))))
            plaf = min(p["brut"], PLAFOND_CNSS)
            tot["reel"] += p["brut"]
            tot["plaf"] += plaf
            tot["cnss"] += p["cnss"]
            tot["amo"] += p["amo"]
            ws.append([rec.get("cnss", ""), rec.get("nom", ""), rec.get("cin", ""),
                       jours, p["brut"], plaf, p["cnss"], p["amo"]])
            for c in range(1, len(cols) + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                if c >= 5:
                    cell.number_format = '# ##0.00 "DH"'
            r += 1

        ws.append(["", "TOTAUX", "", "", round(tot["reel"], 2),
                   round(tot["plaf"], 2), round(tot["cnss"], 2),
                   round(tot["amo"], 2)])
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill, cell.font, cell.border = tot_fill, bold, border
            if c >= 5:
                cell.number_format = '# ##0.00 "DH"'

        widths = [14, 26, 14, 8, 14, 16, 14, 14]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = w
        ws.freeze_panes = "A3"

        try:
            wb.save(out_file)
        except PermissionError:
            messagebox.showerror("الملف محلول", "سدّ الملف فـExcel وعاود حاول.")
            return
        try:
            os.startfile(out_file)
        except (AttributeError, OSError):
            webbrowser.open(out_file.as_uri())
        self.set_status(f"Déclaration CNSS تصدّرات: {out_file.name}")


    def open_graph(self):
        if not self.records:
            messagebox.showinfo("فارغ", "ماكاين حتى خدام باش نرسمو.")
            return
        win = tk.Toplevel(self)
        win.title("Graphique des salaires nets")
        win.configure(bg=COL["surface"])
        win.geometry("760x520")

        tk.Label(win, text="📈  Salaires nets par employé", bg=COL["surface"],
                 fg=COL["brand"], font=(FONT, 14, "bold")).pack(
                     anchor=tk.W, padx=16, pady=(14, 4))

        cv = tk.Canvas(win, bg=COL["surface"], highlightthickness=0)
        cv.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 16))

        data = sorted(((r.get("nom", "?") or "?", compute_payroll(r)["net"])
                       for r in self.records),
                      key=lambda x: x[1], reverse=True)[:15]

        def draw(_event=None):
            cv.delete("all")
            w = cv.winfo_width() or 720
            top, bottom, left, right = 20, 30, 150, 60
            avail_h = cv.winfo_height() - top - bottom
            if not data or avail_h <= 0:
                return
            maxv = max(v for _, v in data) or 1
            n = len(data)
            gap = 8
            bar_h = max(14, (avail_h - gap * n) / n)
            bar_w_max = w - left - right
            for i, (nom, val) in enumerate(data):
                y = top + i * (bar_h + gap)
                color = AVATAR_COLORS[abs(hash(nom)) % len(AVATAR_COLORS)]
                length = bar_w_max * (val / maxv)
                cv.create_text(left - 10, y + bar_h / 2, text=nom[:20], anchor=tk.E,
                               fill=COL["text"], font=(FONT, 9))
                cv.create_rectangle(left, y, left + length, y + bar_h,
                                    fill=color, outline=color)
                cv.create_text(left + length + 6, y + bar_h / 2,
                               text=fmt_money(val), anchor=tk.W,
                               fill=COL["muted"], font=(FONT, 9, "bold"))

        cv.bind("<Configure>", draw)


    def archive_month(self):
        if not self.records:
            messagebox.showinfo("فارغ", "ماكاين حتى خدام باش نأرشيفو.")
            return
        mois = f"{dt.date.today():%Y-%m}"
        existing = self.history.load()
        if mois in existing and not messagebox.askyesno(
                "موجود", f"الشهر {mois} مأرشيف من قبل. تبدلو بالمعطيات الحالية؟"):
            return
        self.history.archive(mois, self.records, compute_payroll)
        messagebox.showinfo("تم", f"تأرشف الشهر {mois} ✓\n"
                                  "دابا تقدر تشوف التطوّر فـ 📉 Évolution.")
        self.set_status(f"الشهر {mois} تأرشف.")

    def open_evolution(self):
        series = self.history.series()
        win = tk.Toplevel(self)
        win.title("Évolution de la masse salariale")
        win.configure(bg=COL["surface"])
        win.geometry("780x520")
        tk.Label(win, text="📉  Évolution de la masse salariale (net)",
                 bg=COL["surface"], fg=COL["brand"], font=(FONT, 14, "bold")).pack(
                     anchor=tk.W, padx=16, pady=(14, 2))
        if not series:
            tk.Label(win, text="ماكاين حتى أرشيف. دير 📦 Archiver باش تبدا تجمع.",
                     bg=COL["surface"], fg=COL["muted"], font=(FONT, 11)).pack(
                         padx=16, pady=24)
            return
        tk.Label(win, text="Astuce : archivez chaque mois pour suivre la tendance.",
                 bg=COL["surface"], fg=COL["muted"], font=(FONT, 9)).pack(
                     anchor=tk.W, padx=16)

        cv = tk.Canvas(win, bg=COL["surface"], highlightthickness=0)
        cv.pack(fill=tk.BOTH, expand=True, padx=16, pady=12)

        def draw(_e=None):
            cv.delete("all")
            w = cv.winfo_width() or 740
            h = cv.winfo_height() or 420
            left, right, top, bottom = 70, 30, 20, 50
            data = [(m, s.get("net", 0)) for m, s in series]
            maxv = max((v for _, v in data), default=1) or 1
            plot_w = w - left - right
            plot_h = h - top - bottom
            cv.create_line(left, top, left, top + plot_h, fill=COL["border"])
            cv.create_line(left, top + plot_h, left + plot_w, top + plot_h,
                           fill=COL["border"])
            n = len(data)
            bw = plot_w / max(n, 1)
            pts = []
            for i, (mois, val) in enumerate(data):
                x = left + bw * (i + 0.5)
                bh = plot_h * (val / maxv)
                y = top + plot_h - bh
                cv.create_rectangle(x - bw * 0.3, y, x + bw * 0.3, top + plot_h,
                                    fill=COL["accent"], outline=COL["accent"])
                cv.create_text(x, top + plot_h + 14, text=mois,
                               fill=COL["muted"], font=(FONT, 8))
                cv.create_text(x, y - 8, text=f"{val/1000:.0f}k",
                               fill=COL["text"], font=(FONT, 8, "bold"))
                pts.append((x, y))
            if len(pts) > 1:
                cv.create_line(pts, fill=COL["brand2"], width=2, smooth=True)

        cv.bind("<Configure>", draw)


    def _conges(self, rec: dict):
        de = parse_date(rec.get("date_embauche"))
        today = dt.date.today()
        if de:
            months = (today.year - de.year) * 12 + (today.month - de.month)
            if today.day < de.day:
                months -= 1
            months = max(0, months)
        else:
            months = 0
        acquis = round(months * LEAVE_PER_MONTH, 1)
        emp_id = rec.get("id")
        pris = 0
        if emp_id:
            pris = self.pointage.count_status(emp_id, "C")
            if hasattr(self, "leaves"):
                pris += self.leaves.approved_annual_days(emp_id)
        return acquis, pris, round(acquis - pris, 1)


    def compute_alerts(self) -> list:
        today = dt.date.today()

        def days_to_anniv(d):
            for yr in (today.year, today.year + 1):
                try:
                    nd = d.replace(year=yr)
                except ValueError:
                    nd = dt.date(yr, 3, 1)
                if nd >= today:
                    return (nd - today).days
            return None

        out = []
        for rec in self.records:
            nom = rec.get("nom", "?") or "?"
            dn = parse_date(rec.get("date_naissance"))
            if dn:
                d = days_to_anniv(dn)
                if d is not None and d <= 7:
                    quand = "aujourd'hui 🎂" if d == 0 else f"dans {d} j"
                    out.append(("🎂", f"Anniversaire de {nom} ({quand})", d))
            de = parse_date(rec.get("date_embauche"))
            if de:
                yrs = today.year - de.year - (
                    1 if (today.month, today.day) < (de.month, de.day) else 0)
                d = days_to_anniv(de)
                if d is not None and d <= 7 and yrs >= 0:
                    quand = "aujourd'hui 🎉" if d == 0 else f"dans {d} j"
                    out.append(("🎉", f"Ancienneté de {nom} ({quand})", d))
            df = parse_date(rec.get("date_fin_contrat"))
            if df:
                delta = (df - today).days
                if delta < 0:
                    out.append(("⚠️", f"Contrat de {nom} expiré (il y a {-delta} j)",
                                -1000 - delta))
                elif delta <= 30:
                    quand = "aujourd'hui" if delta == 0 else f"dans {delta} j"
                    out.append(("📄", f"Fin de contrat de {nom} ({quand})", delta))
        out.sort(key=lambda x: x[2])
        return out

    def update_alert_badge(self):
        if not hasattr(self, "alert_btn"):
            return
        n = len(self.compute_alerts())
        self.alert_btn.config(text=f"🔔  Alertes ({n})" if n else "🔔  Alertes")

    def open_alerts(self):
        alerts = self.compute_alerts()
        win = tk.Toplevel(self)
        win.title("Alertes RH")
        win.configure(bg=COL["surface"])
        win.geometry("480x420")
        tk.Label(win, text="🔔  Alertes RH", bg=COL["surface"], fg=COL["brand"],
                 font=(FONT, 14, "bold")).pack(anchor=tk.W, padx=16, pady=(14, 8))
        if not alerts:
            tk.Label(win, text="Aucune alerte 🎉  — tout est à jour.",
                     bg=COL["surface"], fg=COL["muted"],
                     font=(FONT, 11)).pack(padx=16, pady=20)
            return
        wrap = tk.Frame(win, bg=COL["surface"])
        wrap.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 16))
        for icon, text, _ in alerts:
            row = tk.Frame(wrap, bg=COL["bg"])
            row.pack(fill=tk.X, pady=3)
            tk.Label(row, text=icon, bg=COL["bg"],
                     font=(FONT, 14)).pack(side=tk.LEFT, padx=8, pady=6)
            tk.Label(row, text=text, bg=COL["bg"], fg=COL["text"],
                     font=(FONT, 10), anchor=tk.W, justify=tk.LEFT).pack(
                         side=tk.LEFT, padx=(0, 8))


    def open_year_pointage(self):
        if self.current_index is None:
            messagebox.showinfo("اختار خدام", "اختار شي خدام الأول.")
            return
        rec = self.records[self.current_index]
        emp_id = rec.get("id")
        fr_months = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                     "Juillet", "Août", "Septembre", "Octobre", "Novembre",
                     "Décembre"]
        st = {"y": dt.date.today().year}

        win = tk.Toplevel(self)
        win.title(f"Calendrier annuel — {rec.get('nom', '')}")
        win.configure(bg=COL["surface"])
        win.geometry("560x520")
        tk.Label(win, text=f"🗓  Présence annuelle — {rec.get('nom', '')}",
                 bg=COL["surface"], fg=COL["brand"], font=(FONT, 13, "bold")).pack(
                     anchor=tk.W, padx=16, pady=(14, 4))

        nav = tk.Frame(win, bg=COL["surface"])
        nav.pack(padx=16, pady=2)
        year_lbl = tk.Label(nav, bg=COL["surface"], fg=COL["text"],
                            font=(FONT, 12, "bold"), width=8)
        tb.Button(nav, text="◀", bootstyle="secondary-outline",
                   command=lambda: shift(-1)).pack(side=tk.LEFT)
        year_lbl.pack(side=tk.LEFT, padx=8)
        tb.Button(nav, text="▶", bootstyle="secondary-outline",
                   command=lambda: shift(1)).pack(side=tk.LEFT)

        cols = ("mois", "P", "A", "C", "R")
        tree = ttk.Treeview(win, columns=cols, show="headings", height=14)
        for c, lab, w in (("mois", "Mois", 130), ("P", "Présents", 80),
                          ("A", "Absents", 80), ("C", "Congés", 80),
                          ("R", "Repos", 80)):
            tree.heading(c, text=lab)
            tree.column(c, width=w, anchor=tk.CENTER if c != "mois" else tk.W)
        tree.pack(fill=tk.BOTH, expand=True, padx=16, pady=8)

        def render():
            tree.delete(*tree.get_children())
            year_lbl.config(text=str(st["y"]))
            data = self.pointage.all_for(emp_id)
            tot = {"P": 0, "A": 0, "C": 0, "R": 0}
            for m in range(1, 13):
                key = f"{st['y']}-{m:02d}"
                sts = data.get(key, [])
                cnt = {c: sts.count(c) for c in ("P", "A", "C", "R")}
                for c in tot:
                    tot[c] += cnt[c]
                vals = (fr_months[m], cnt["P"] or "—", cnt["A"] or "—",
                        cnt["C"] or "—", cnt["R"] or "—")
                tree.insert("", tk.END, values=vals)
            tree.insert("", tk.END, values=("TOTAL", tot["P"], tot["A"],
                                            tot["C"], tot["R"]), tags=("tot",))
            tree.tag_configure("tot", background=COL["net_bg"])

        def shift(d):
            st["y"] += d
            render()

        tb.Button(win, text="Fermer", bootstyle="secondary-outline",
                   command=win.destroy).pack(pady=(0, 14))
        render()


    def open_pointage(self):
        if self.current_index is None:
            messagebox.showinfo("اختار خدام", "اختار شي خدام الأول باش تسجل الحضور.")
            return
        rec = self.records[self.current_index]
        emp_id = rec.get("id")
        today = dt.date.today()
        st = {"y": today.year, "m": today.month, "statuses": [], "btns": {}}
        fr_months = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                     "Juillet", "Août", "Septembre", "Octobre", "Novembre",
                     "Décembre"]

        win = tk.Toplevel(self)
        win.title(f"Pointage — {rec.get('nom','')}")
        win.configure(bg=COL["surface"])
        win.geometry("560x600")

        top = tk.Frame(win, bg=COL["surface"])
        top.pack(fill=tk.X, padx=16, pady=(14, 4))
        tk.Label(top, text=f"📅  {rec.get('nom','')}", bg=COL["surface"],
                 fg=COL["brand"], font=(FONT, 13, "bold")).pack(side=tk.LEFT)

        nav = tk.Frame(win, bg=COL["surface"])
        nav.pack(fill=tk.X, padx=16)
        month_lbl = tk.Label(nav, text="", bg=COL["surface"], fg=COL["text"],
                             font=(FONT, 12, "bold"), width=18)
        tb.Button(nav, text="◀", bootstyle="secondary-outline",
                   command=lambda: shift(-1)).pack(side=tk.LEFT)
        month_lbl.pack(side=tk.LEFT, padx=8)
        tb.Button(nav, text="▶", bootstyle="secondary-outline",
                   command=lambda: shift(1)).pack(side=tk.LEFT)

        legend = tk.Frame(win, bg=COL["surface"])
        legend.pack(fill=tk.X, padx=16, pady=(6, 2))
        for code in POINTAGE_STATUSES:
            tk.Label(legend, text=f"  {code} ", bg=POINTAGE_COLORS[code],
                     fg="white", font=(FONT, 8, "bold")).pack(side=tk.LEFT, padx=(0, 2))
            tk.Label(legend, text=POINTAGE_LABELS[code], bg=COL["surface"],
                     fg=COL["muted"], font=(FONT, 8)).pack(side=tk.LEFT, padx=(0, 10))

        grid = tk.Frame(win, bg=COL["surface"])
        grid.pack(padx=16, pady=8)

        summary = tk.Label(win, bg=COL["surface"], fg=COL["text"],
                           font=(FONT, 10, "bold"))
        summary.pack(pady=(4, 8))

        def update_summary():
            cnt = {c: st["statuses"].count(c) for c in POINTAGE_STATUSES}
            summary.config(
                text=f"Présents: {cnt['P']}   |   Absents: {cnt['A']}   |   "
                     f"Congés: {cnt['C']}   |   Repos: {cnt['R']}")

        def paint(day):
            b = st["btns"][day]
            code = st["statuses"][day - 1]
            b.config(text=f"{day}\n{code}", bg=POINTAGE_COLORS[code], fg="white")

        def cycle(day):
            cur = st["statuses"][day - 1]
            nxt = POINTAGE_STATUSES[
                (POINTAGE_STATUSES.index(cur) + 1) % len(POINTAGE_STATUSES)]
            st["statuses"][day - 1] = nxt
            paint(day)
            update_summary()

        def render():
            for w in grid.winfo_children():
                w.destroy()
            st["btns"] = {}
            y, m = st["y"], st["m"]
            month_lbl.config(text=f"{fr_months[m]} {y}")
            mois = f"{y}-{m:02d}"
            ndays = calendar.monthrange(y, m)[1]
            loaded = self.pointage.get(emp_id, mois)
            st["statuses"] = (list(loaded) if len(loaded) == ndays
                              else default_pointage(y, m))
            for i, dname in enumerate(["Lun", "Mar", "Mer", "Jeu", "Ven",
                                       "Sam", "Dim"]):
                tk.Label(grid, text=dname, bg=COL["surface"], fg=COL["muted"],
                         font=(FONT, 9, "bold"), width=5).grid(row=0, column=i,
                                                               pady=(0, 4))
            for r, week in enumerate(calendar.monthcalendar(y, m), start=1):
                for c, day in enumerate(week):
                    if day == 0:
                        continue
                    b = tk.Button(grid, width=5, height=2, bd=0, cursor="hand2",
                                  font=(FONT, 8, "bold"),
                                  command=lambda d=day: cycle(d))
                    b.grid(row=r, column=c, padx=2, pady=2)
                    st["btns"][day] = b
                    paint(day)
            update_summary()

        def shift(delta):
            m = st["m"] + delta
            y = st["y"]
            if m < 1:
                m, y = 12, y - 1
            elif m > 12:
                m, y = 1, y + 1
            st["m"], st["y"] = m, y
            render()

        def save():
            mois = f"{st['y']}-{st['m']:02d}"
            self.pointage.set(emp_id, mois, st["statuses"])
            absc = st["statuses"].count("A")
            self.vars["jours_absence"].set(str(absc))
            self.save_record()
            messagebox.showinfo(
                "تسجل", f"تسجل الحضور ديال {mois} ✓\n"
                        f"أيام الغياب ({absc}) تطبقات على الأجرة.", parent=win)
            win.destroy()

        btns = tk.Frame(win, bg=COL["surface"])
        btns.pack(pady=(4, 14))
        tb.Button(btns, text="💾  Enregistrer + appliquer", bootstyle="success",
                   command=save).pack(side=tk.LEFT, padx=6)
        tb.Button(btns, text="Fermer", bootstyle="secondary-outline",
                   command=win.destroy).pack(side=tk.LEFT, padx=6)

        render()


    def export_pdf(self):
        if not HAS_FPDF:
            messagebox.showinfo("fpdf2 ناقصة",
                                "باش تصدر PDF ثبت:\n\n    pip install fpdf2")
            return
        rec = self.collect_form()
        if not rec.get("nom"):
            messagebox.showwarning("ناقص", "اختار شي خدام أو عمر الفيش الأول.")
            return
        rec["anciennete"] = compute_anciennete(rec.get("date_embauche"))
        try:
            path = self._bulletin_pdf(rec)
        except Exception as exc:
            messagebox.showerror("خطأ PDF", str(exc))
            return
        try:
            os.startfile(path)
        except (AttributeError, OSError):
            webbrowser.open(Path(path).as_uri())
        self.set_status(f"بولتان PDF تصدّر: {path.name}")

    def _setup_pdf_font(self, pdf) -> str:
        font = "helvetica"
        try:
            ar = Path(r"C:\Windows\Fonts\arial.ttf")
            arb = Path(r"C:\Windows\Fonts\arialbd.ttf")
            if ar.exists() and arb.exists():
                pdf.add_font("uni", "", str(ar))
                pdf.add_font("uni", "B", str(arb))
                font = "uni"
        except Exception:
            font = "helvetica"
        return font

    def _bulletin_pdf(self, rec: dict):
        pdf = FPDF(format="A4")
        pdf.set_auto_page_break(True, 15)
        font = self._setup_pdf_font(pdf)
        pdf.add_page()
        self._render_bulletin_page(pdf, rec, font)
        out_dir = self.store.path.parent / "fiches"
        out_dir.mkdir(exist_ok=True)
        safe = "".join(c for c in str(rec.get("nom")) if c.isalnum()
                       or c in " _-").strip().replace(" ", "_")
        out_file = out_dir / f"bulletin_{rec.get('id','')}_{safe}.pdf"
        pdf.output(str(out_file))
        return out_file

    def export_all_pdf(self):
        if not HAS_FPDF:
            messagebox.showinfo("fpdf2 ناقصة",
                                "باش تصدر PDF ثبت:\n\n    pip install fpdf2")
            return
        if not self.records:
            messagebox.showinfo("فارغ", "ماكاين حتى خدام.")
            return
        pdf = FPDF(format="A4")
        pdf.set_auto_page_break(True, 15)
        font = self._setup_pdf_font(pdf)
        for rec in self.records:
            rec = dict(rec)
            rec["anciennete"] = compute_anciennete(rec.get("date_embauche"))
            pdf.add_page()
            self._render_bulletin_page(pdf, rec, font)
        out_dir = self.store.path.parent / "fiches"
        out_dir.mkdir(exist_ok=True)
        out_file = out_dir / f"bulletins_{dt.date.today():%Y_%m}.pdf"
        try:
            pdf.output(str(out_file))
        except PermissionError:
            messagebox.showerror("الملف محلول", "سدّ ملف PDF وعاود حاول.")
            return
        try:
            os.startfile(out_file)
        except (AttributeError, OSError):
            webbrowser.open(out_file.as_uri())
        self.set_status(
            f"{len(self.records)} بولتانات تصدّرو فـ {out_file.name}")

    def _render_bulletin_page(self, pdf, rec: dict, font: str):
        p = compute_payroll(rec)
        _, _, solde = self._conges(rec)
        brand = (15, 61, 46)
        brand2 = (21, 128, 61)
        netbg = (234, 250, 240)
        red = (185, 28, 28)

        def t(s):
            s = str(s)
            if font == "uni":
                return s
            return s.encode("latin-1", "replace").decode("latin-1")

        def setf(size, style=""):
            pdf.set_font(font, style, size)

        pdf.set_fill_color(*brand)
        pdf.rect(0, 0, 210, 28, "F")
        pdf.set_text_color(255, 255, 255)
        setf(16, "B")
        pdf.set_xy(12, 7)
        pdf.cell(120, 8, t("Bulletin de paie"))
        setf(10, "")
        pdf.set_xy(12, 16)
        ref = f"BUL-{dt.date.today():%Y%m}-{rec.get('id', '')}"
        pdf.cell(120, 6, t(f"Période : {dt.date.today():%m/%Y}   |   Réf : {ref}"))
        logo_ok = False
        if LOGO_PATH and os.path.exists(LOGO_PATH):
            try:
                pdf.image(LOGO_PATH, x=176, y=5, h=18)
                logo_ok = True
            except Exception:
                logo_ok = False
        setf(13, "B")
        pdf.set_xy(112, 11)
        pdf.cell(60 if logo_ok else 78, 8, t(COMPANY_NAME), align="R")

        photo = str(rec.get("photo", "")).strip()
        if photo and os.path.exists(photo):
            try:
                pdf.image(photo, x=170, y=32, w=26, h=26)
            except Exception:
                pass

        pdf.set_text_color(30, 41, 59)
        pdf.set_xy(12, 34)
        setf(15, "B")
        pdf.cell(150, 10, t(rec.get("nom", "")), ln=1)
        pdf.ln(2)

        info = [("Matricule", rec.get("id", "")), ("CIN", rec.get("cin", "")),
                ("Poste", rec.get("poste", "")), ("N° CNSS", rec.get("cnss", "")),
                ("Date d'embauche", rec.get("date_embauche", "")),
                ("Ancienneté", rec.get("anciennete", "")),
                ("Type de contrat", rec.get("type_contrat", "")),
                ("Personnes à charge", rec.get("personnes_charge", "") or "0"),
                ("Solde congés", f"{solde} j")]
        pdf.set_x(12)
        for i, (k, v) in enumerate(info):
            setf(9, "B")
            pdf.set_fill_color(241, 245, 249)
            pdf.cell(45, 7, t(k), border=1, fill=True)
            setf(9, "")
            pdf.cell(48, 7, t(v), border=1, ln=(1 if i % 2 else 0))
            if i % 2 == 0:
                pdf.set_x(105)
        if len(info) % 2 == 1:
            pdf.ln(7)
        pdf.ln(4)

        def hcell(txt, w, align="L"):
            setf(10, "B")
            pdf.set_fill_color(*brand2)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(w, 8, t(txt), border=1, fill=True, align=align)

        pdf.set_x(12)
        hcell("Rubrique", 90)
        hcell("Gains", 48, "R")
        hcell("Retenues", 48, "R")
        pdf.ln(8)
        pdf.set_text_color(30, 41, 59)

        rows = [("Salaire de base", fmt_money(p["base"]), ""),
                ("Primes / indemnités", fmt_money(p["primes"]), ""),
                ("CNSS (4.48%)", "", fmt_money(p["cnss"])),
                ("AMO (2.26%)", "", fmt_money(p["amo"])),
                ("IR (impôt)", "", fmt_money(p["ir"])),
                ("Retenue absence", "", fmt_money(p["ded_absence"])),
                ("Autres retenues", "", fmt_money(p["autres"])),
                ("Retenue avance", "", fmt_money(p["avance"]))]
        for label, gain, ret in rows:
            pdf.set_x(12)
            setf(9, "")
            pdf.cell(90, 7, t(label), border=1)
            pdf.cell(48, 7, t(gain), border=1, align="R")
            pdf.set_text_color(*red)
            pdf.cell(48, 7, t(ret), border=1, align="R", ln=1)
            pdf.set_text_color(30, 41, 59)

        pdf.set_x(12)
        setf(10, "B")
        pdf.set_fill_color(241, 245, 249)
        pdf.cell(90, 8, t("Totaux"), border=1, fill=True)
        pdf.cell(48, 8, t(fmt_money(p["brut"])), border=1, align="R", fill=True)
        pdf.set_text_color(*red)
        pdf.cell(48, 8, t(fmt_money(p["total_retenues"])), border=1,
                 align="R", fill=True, ln=1)
        pdf.set_text_color(30, 41, 59)
        pdf.ln(4)

        pdf.set_x(12)
        setf(14, "B")
        pdf.set_fill_color(*netbg)
        pdf.set_text_color(*brand2)
        pdf.cell(186, 12, t(f"  Net à payer : {fmt_money(p['net'])}"),
                 border=1, fill=True, align="R", ln=1)

        pdf.set_text_color(148, 163, 184)
        setf(8, "")
        pdf.ln(4)
        pdf.set_x(12)
        pdf.cell(186, 5, t(f"Édité le {dt.date.today().isoformat()} — "
                           f"{COMPANY_NAME} — Document indicatif."), align="C")


    def import_csv(self):
        if not self._require_admin():
            return
        path = filedialog.askopenfilename(
            title="استورد لائحة الخدامة (CSV)",
            filetypes=[("CSV", "*.csv"), ("Tous", "*.*")])
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8-sig", newline="") as fh:
                reader = csv.DictReader(fh)
                if not reader.fieldnames:
                    messagebox.showwarning("فارغ", "الملف ماعندو رؤوس أعمدة.")
                    return
                colmap = {}
                for col in reader.fieldnames:
                    c = (col or "").strip()
                    if c in HEADER_TO_KEY:
                        colmap[col] = HEADER_TO_KEY[c]
                    elif c in FIELD_BY_KEY:
                        colmap[col] = c
                if "nom" not in colmap.values():
                    messagebox.showwarning(
                        "ناقص", "خاص الملف يحتوي على عمود 'Nom complet' أو 'nom'.")
                    return
                rows = list(reader)
        except Exception as exc:
            messagebox.showerror("خطأ القراءة", str(exc))
            return

        next_id = self.next_id()
        added = 0
        for row in rows:
            rec = {k: "" for k in FIELD_BY_KEY}
            for col, key in colmap.items():
                rec[key] = str(row.get(col, "") or "").strip()
            if not rec.get("nom"):
                continue
            if not str(rec.get("id")).strip():
                rec["id"] = str(next_id)
                next_id += 1
            p = compute_payroll(rec)
            rec["salaire_brut"] = f"{p['brut']:.2f}"
            rec["ret_cnss"] = f"{p['cnss']:.2f}"
            rec["ret_amo"] = f"{p['amo']:.2f}"
            rec["ret_ir"] = f"{p['ir']:.2f}"
            rec["anciennete"] = compute_anciennete(rec.get("date_embauche"))
            rec["salaire_net"] = f"{p['net']:.2f}"
            self.records.append(rec)
            added += 1

        if not added:
            messagebox.showinfo("والو", "حتى صف صالح ماتزاد (خاص على الأقل الاسم).")
            return
        try:
            self.store.save_all(self.records)
        except Exception as exc:
            messagebox.showerror("خطأ الحفظ", str(exc))
            return
        self.reload()
        messagebox.showinfo("تم", f"تزادو {added} خدّام من CSV ✓")
        self.set_status(f"استيراد CSV: {added} خدّام.")

    def export_csv(self):
        if not self.records:
            messagebox.showinfo("فارغ", "ماكاين حتى خدام.")
            return
        path = filedialog.asksaveasfilename(
            title="صدّر لائحة الخدامة (CSV)", defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialdir=str(self.store.path.parent), initialfile="employes_export.csv")
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8-sig", newline="") as fh:
                writer = csv.writer(fh)
                writer.writerow(HEADERS)
                for rec in self.records:
                    writer.writerow([rec.get(HEADER_TO_KEY[h], "") for h in HEADERS])
        except Exception as exc:
            messagebox.showerror("خطأ", str(exc))
            return
        self.set_status(f"تصدّر CSV: {Path(path).name}")


    def open_advances(self):
        if self.current_index is None:
            messagebox.showinfo("اختار خدام", "اختار شي خدام الأول.")
            return
        rec = self.records[self.current_index]
        emp_id = str(rec.get("id"))

        win = tk.Toplevel(self)
        win.title(f"Avances — {rec.get('nom', '')}")
        win.configure(bg=COL["surface"])
        win.geometry("560x480")
        tk.Label(win, text=f"💳  Avances — {rec.get('nom', '')}", bg=COL["surface"],
                 fg=COL["brand"], font=(FONT, 13, "bold")).pack(
                     anchor=tk.W, padx=16, pady=(14, 2))
        solde_lbl = tk.Label(win, bg=COL["surface"], fg=COL["brand2"],
                             font=(FONT, 11, "bold"))
        solde_lbl.pack(anchor=tk.W, padx=16)

        cols = ("date", "montant", "mensualite", "solde")
        tree = ttk.Treeview(win, columns=cols, show="headings", height=8)
        for c, lab, w in (("date", "Date", 100), ("montant", "Montant", 110),
                          ("mensualite", "Mensualité", 110), ("solde", "Solde", 110)):
            tree.heading(c, text=lab)
            tree.column(c, width=w, anchor=tk.E if c != "date" else tk.CENTER)
        tree.pack(fill=tk.BOTH, expand=True, padx=16, pady=8)

        def refresh():
            tree.delete(*tree.get_children())
            for i, a in enumerate(self.advances.list(emp_id)):
                tree.insert("", tk.END, iid=str(i), values=(
                    a.get("date", ""), fmt_money(a.get("montant", 0)),
                    fmt_money(a.get("mensualite", 0)), fmt_money(a.get("solde", 0))))
            solde_lbl.config(
                text=f"Solde total dû : {fmt_money(self.advances.solde_total(emp_id))}")

        form = tk.Frame(win, bg=COL["surface"])
        form.pack(fill=tk.X, padx=16, pady=(0, 6))
        tk.Label(form, text="Montant :", bg=COL["surface"], fg=COL["text"],
                 font=(FONT, 9)).grid(row=0, column=0, padx=2)
        v_montant = tk.StringVar()
        ttk.Entry(form, textvariable=v_montant, width=10).grid(row=0, column=1, padx=2)
        tk.Label(form, text="Nb mois :", bg=COL["surface"], fg=COL["text"],
                 font=(FONT, 9)).grid(row=0, column=2, padx=2)
        v_mois = tk.StringVar(value="1")
        ttk.Entry(form, textvariable=v_mois, width=6).grid(row=0, column=3, padx=2)

        def add_av():
            m = to_float(v_montant.get())
            n = int(to_float(v_mois.get())) or 1
            if m <= 0:
                messagebox.showwarning("خطأ", "دخّل مبلغ صحيح.", parent=win)
                return
            self.advances.add(emp_id, m, round(m / n, 2))
            v_montant.set("")
            refresh()

        tb.Button(form, text="➕  Ajouter une avance", bootstyle="success",
                   command=add_av).grid(row=0, column=4, padx=8)

        def supprimer():
            sel = tree.selection()
            if not sel:
                return
            advances = self.advances.list(emp_id)
            idx = int(sel[0])
            if 0 <= idx < len(advances) and messagebox.askyesno(
                    "مسح", "تمسح هاد السلفة؟", parent=win):
                del advances[idx]
                self.advances.set(emp_id, advances)
                refresh()

        def prelever():
            advances = self.advances.list(emp_id)
            total = 0.0
            for a in advances:
                solde = a.get("solde", 0)
                if solde <= 0:
                    continue
                pay = min(a.get("mensualite", 0), solde)
                a["solde"] = round(solde - pay, 2)
                total += pay
            if total <= 0:
                messagebox.showinfo("والو", "ماكاين حتى سلفة باقية.", parent=win)
                return
            self.advances.set(emp_id, advances)
            self.vars["retenue_avance"].set(f"{round(total, 2):.2f}")
            self.save_record()
            refresh()
            messagebox.showinfo(
                "تم", f"تخصمات منسوالية {fmt_money(total)} من أجرة هاد الشهر ✓",
                parent=win)

        bb = tk.Frame(win, bg=COL["surface"])
        bb.pack(fill=tk.X, padx=16, pady=(0, 14))
        tb.Button(bb, text="💸  Prélever la mensualité (ce mois)",
                   bootstyle="success", command=prelever).pack(side=tk.LEFT)
        tb.Button(bb, text="🗑  Supprimer", bootstyle="danger",
                   command=supprimer).pack(side=tk.LEFT, padx=6)
        tb.Button(bb, text="Fermer", bootstyle="secondary-outline",
                   command=win.destroy).pack(side=tk.RIGHT)
        refresh()

    def open_leaves(self):
        if self.current_index is None:
            messagebox.showinfo("اختار خدام", "اختار شي خدام الأول.")
            return
        rec = self.records[self.current_index]
        emp_id = rec.get("id")

        win = tk.Toplevel(self)
        win.title(f"Congés — {rec.get('nom', '')}")
        win.configure(bg=COL["surface"])
        win.geometry("640x520")
        tk.Label(win, text=f"🏖  Congés — {rec.get('nom', '')}", bg=COL["surface"],
                 fg=COL["brand"], font=(FONT, 13, "bold")).pack(
                     anchor=tk.W, padx=16, pady=(14, 2))
        bal_lbl = tk.Label(win, bg=COL["surface"], fg=COL["brand2"],
                           font=(FONT, 11, "bold"))
        bal_lbl.pack(anchor=tk.W, padx=16)

        cols = ("type", "start", "end", "days", "status")
        tree = ttk.Treeview(win, columns=cols, show="headings", height=9)
        for c, lab, w in (("type", "Type", 100), ("start", "Du", 110),
                          ("end", "Au", 110), ("days", "Jours", 60),
                          ("status", "Statut", 110)):
            tree.heading(c, text=lab)
            tree.column(c, width=w, anchor=tk.W)
        tree.pack(fill=tk.BOTH, expand=True, padx=16, pady=8)

        def refresh():
            tree.delete(*tree.get_children())
            for i, r in enumerate(self.leaves.list(emp_id)):
                tree.insert("", tk.END, iid=str(i), values=(
                    r.get("type", ""), r.get("start", ""), r.get("end", ""),
                    r.get("days", ""), r.get("status", "")))
            acq, pris, solde = self._conges(rec)
            bal_lbl.config(text=f"Solde : acquis {acq} j  •  pris {pris} j  "
                                f"•  restant {solde} j")

        form = tk.Frame(win, bg=COL["surface"])
        form.pack(fill=tk.X, padx=16, pady=(0, 4))
        v_type = tk.StringVar(value="Annuel")
        v_start = tk.StringVar()
        v_end = tk.StringVar()
        ttk.Combobox(form, textvariable=v_type, values=LEAVE_TYPES,
                     state="readonly", width=10).grid(row=0, column=0, padx=2)
        tk.Label(form, text="Du", bg=COL["surface"], fg=COL["muted"],
                 font=(FONT, 9)).grid(row=0, column=1, padx=(6, 2))
        ttk.Entry(form, textvariable=v_start, width=12).grid(row=0, column=2, padx=2)
        tk.Label(form, text="Au", bg=COL["surface"], fg=COL["muted"],
                 font=(FONT, 9)).grid(row=0, column=3, padx=(6, 2))
        ttk.Entry(form, textvariable=v_end, width=12).grid(row=0, column=4, padx=2)

        def add_req():
            d1, d2 = parse_date(v_start.get()), parse_date(v_end.get())
            if not d1 or not d2 or d2 < d1:
                messagebox.showwarning(
                    "Dates", "Dates invalides (AAAA-MM-JJ, fin ≥ début).",
                    parent=win)
                return
            days = (d2 - d1).days + 1
            self.leaves.add(emp_id, v_type.get(), d1.isoformat(),
                            d2.isoformat(), days)
            self.audit.log(getattr(self, "role", "admin"), "Demande congé",
                           f"{rec.get('nom')} {v_type.get()} {days}j")
            v_start.set(""); v_end.set("")
            refresh()

        tb.Button(form, text="➕  Demander", bootstyle="info",
                  command=add_req).grid(row=0, column=5, padx=8)

        def set_status(new):
            sel = tree.selection()
            if not sel:
                return
            reqs = self.leaves.list(emp_id)
            idx = int(sel[0])
            reqs[idx]["status"] = new
            self.leaves.set_all(emp_id, reqs)
            self.audit.log(getattr(self, "role", "admin"), f"Congé {new}",
                           f"{rec.get('nom')} {reqs[idx].get('type')} "
                           f"{reqs[idx].get('days')}j")
            refresh()
            self.update_calc()

        bb = tk.Frame(win, bg=COL["surface"])
        bb.pack(fill=tk.X, padx=16, pady=(0, 14))
        tb.Button(bb, text="✓  Approuver", bootstyle="success",
                  command=lambda: set_status("Approuvé")).pack(side=tk.LEFT)
        tb.Button(bb, text="✗  Refuser", bootstyle="danger",
                  command=lambda: set_status("Refusé")).pack(side=tk.LEFT, padx=6)
        tb.Button(bb, text="Fermer", bootstyle="secondary-outline",
                  command=win.destroy).pack(side=tk.RIGHT)
        refresh()

    def open_hours(self):
        if self.current_index is None:
            messagebox.showinfo("اختار خدام", "اختار شي خدام الأول.")
            return
        rec = self.records[self.current_index]
        emp_id = rec.get("id")
        today = dt.date.today()
        st = {"y": today.year, "m": today.month, "sup": 0.0}
        fr_months = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                     "Juillet", "Août", "Septembre", "Octobre", "Novembre",
                     "Décembre"]

        win = tk.Toplevel(self)
        win.title(f"Heures — {rec.get('nom', '')}")
        win.configure(bg=COL["surface"])
        win.geometry("620x600")
        tk.Label(win, text=f"🕐  Heures (entrée/sortie) — {rec.get('nom', '')}",
                 bg=COL["surface"], fg=COL["brand"], font=(FONT, 13, "bold")).pack(
                     anchor=tk.W, padx=16, pady=(14, 2))
        tk.Label(win, text=f"Début standard {STD_START} · journée {STD_DAILY_HOURS}h "
                 f"· heures sup ×{OVERTIME_RATE}", bg=COL["surface"],
                 fg=COL["muted"], font=(FONT, 8)).pack(anchor=tk.W, padx=16)

        nav = tk.Frame(win, bg=COL["surface"])
        nav.pack(padx=16, pady=4)
        month_lbl = tk.Label(nav, bg=COL["surface"], fg=COL["text"],
                             font=(FONT, 12, "bold"), width=16)
        tb.Button(nav, text="◀", bootstyle="secondary-outline",
                  command=lambda: shift(-1)).pack(side=tk.LEFT)
        month_lbl.pack(side=tk.LEFT, padx=8)
        tb.Button(nav, text="▶", bootstyle="secondary-outline",
                  command=lambda: shift(1)).pack(side=tk.LEFT)

        cols = ("jour", "in", "out", "h", "ret", "sup")
        tree = ttk.Treeview(win, columns=cols, show="headings", height=12)
        for c, lab, w in (("jour", "Jour", 80), ("in", "Entrée", 80),
                          ("out", "Sortie", 80), ("h", "Heures", 70),
                          ("ret", "Retard(min)", 90), ("sup", "Sup(h)", 70)):
            tree.heading(c, text=lab)
            tree.column(c, width=w, anchor=tk.CENTER)
        tree.pack(fill=tk.BOTH, expand=True, padx=16, pady=6)

        totals = tk.Label(win, bg=COL["surface"], fg=COL["brand2"],
                          font=(FONT, 10, "bold"))
        totals.pack(anchor=tk.W, padx=16)

        form = tk.Frame(win, bg=COL["surface"])
        form.pack(fill=tk.X, padx=16, pady=6)
        v_day = tk.StringVar(value=str(today.day))
        v_in = tk.StringVar()
        v_out = tk.StringVar()
        for i, (lab, var, wdt) in enumerate(
                (("Jour", v_day, 5), ("Entrée", v_in, 7), ("Sortie", v_out, 7))):
            tk.Label(form, text=lab, bg=COL["surface"], fg=COL["muted"],
                     font=(FONT, 9)).grid(row=0, column=i * 2, padx=(6, 2))
            ttk.Entry(form, textvariable=var, width=wdt).grid(
                row=0, column=i * 2 + 1, padx=2)

        def render():
            tree.delete(*tree.get_children())
            y, m = st["y"], st["m"]
            month_lbl.config(text=f"{fr_months[m]} {y}")
            mois = f"{y}-{m:02d}"
            data = self.hours.get_month(emp_id, mois)
            ndays = calendar.monthrange(y, m)[1]
            th = tr = ts = 0.0
            for d in range(1, ndays + 1):
                day = data.get(str(d), {})
                ti, to = day.get("in", ""), day.get("out", "")
                h, ret, sup = compute_day_hours(ti, to)
                th += h; tr += ret; ts += sup
                tree.insert("", tk.END, iid=str(d), values=(
                    f"{d:02d}/{m:02d}", ti or "—", to or "—",
                    h or "—", ret or "—", sup or "—"))
            st["sup"] = round(ts, 2)
            totals.config(text=f"Total : {round(th,2)} h   •   Retard : "
                               f"{int(tr)} min   •   Heures sup : {round(ts,2)} h")

        def shift(delta):
            m = st["m"] + delta
            y = st["y"]
            if m < 1:
                m, y = 12, y - 1
            elif m > 12:
                m, y = 1, y + 1
            st["m"], st["y"] = m, y
            render()

        def fill_from_row(_e=None):
            sel = tree.selection()
            if not sel:
                return
            d = int(sel[0])
            mois = f"{st['y']}-{st['m']:02d}"
            day = self.hours.get_month(emp_id, mois).get(str(d), {})
            v_day.set(str(d)); v_in.set(day.get("in", "")); v_out.set(day.get("out", ""))

        tree.bind("<Double-1>", fill_from_row)

        def save_day():
            try:
                d = int(v_day.get())
            except ValueError:
                messagebox.showwarning("Jour", "Numéro de jour invalide.", parent=win)
                return
            ndays = calendar.monthrange(st["y"], st["m"])[1]
            if not (1 <= d <= ndays):
                messagebox.showwarning("Jour", f"Jour entre 1 et {ndays}.",
                                       parent=win)
                return
            ti, to = v_in.get().strip(), v_out.get().strip()
            if ti and parse_time(ti) is None or to and parse_time(to) is None:
                messagebox.showwarning("Heure", "Format heure invalide (HH:MM).",
                                       parent=win)
                return
            mois = f"{st['y']}-{st['m']:02d}"
            self.hours.set_day(emp_id, mois, d, ti, to)
            v_in.set(""); v_out.set("")
            render()

        def apply_overtime():
            base = to_float(rec.get("salaire_base"))
            hourly = base / (JOURS_OUVRABLES * STD_DAILY_HOURS) \
                if JOURS_OUVRABLES and STD_DAILY_HOURS else 0
            amount = round(st["sup"] * hourly * OVERTIME_RATE, 2)
            if amount <= 0:
                messagebox.showinfo("Heures sup", "Aucune heure sup ce mois.",
                                    parent=win)
                return
            if not messagebox.askyesno(
                    "Heures sup → Primes",
                    f"Ajouter {fmt_money(amount)} ({st['sup']} h sup) aux primes ?",
                    parent=win):
                return
            cur = to_float(self.vars["primes"].get())
            self.vars["primes"].set(f"{cur + amount:.2f}")
            self.save_record()
            win.destroy()

        tb.Button(form, text="💾  Jour", bootstyle="success",
                  command=save_day).grid(row=0, column=6, padx=8)

        bb = tk.Frame(win, bg=COL["surface"])
        bb.pack(fill=tk.X, padx=16, pady=(0, 14))
        tb.Button(bb, text="➕  Heures sup → Primes", bootstyle="info",
                  command=apply_overtime).pack(side=tk.LEFT)
        tb.Button(bb, text="Fermer", bootstyle="secondary-outline",
                  command=win.destroy).pack(side=tk.RIGHT)
        render()

    def open_documents(self):
        if self.current_index is None:
            messagebox.showinfo("اختار خدام", "اختار شي خدام الأول.")
            return
        rec = self.records[self.current_index]
        emp_id = str(rec.get("id"))
        doc_dir = self.store.path.parent / "documents" / emp_id
        doc_dir.mkdir(parents=True, exist_ok=True)

        win = tk.Toplevel(self)
        win.title(f"Documents — {rec.get('nom', '')}")
        win.configure(bg=COL["surface"])
        win.geometry("540x460")
        tk.Label(win, text=f"📎  Documents — {rec.get('nom', '')}",
                 bg=COL["surface"], fg=COL["brand"], font=(FONT, 13, "bold")).pack(
                     anchor=tk.W, padx=16, pady=(14, 2))
        count_lbl = tk.Label(win, bg=COL["surface"], fg=COL["muted"],
                             font=(FONT, 9))
        count_lbl.pack(anchor=tk.W, padx=16)

        wrap = tk.Frame(win, bg=COL["surface"])
        wrap.pack(fill=tk.BOTH, expand=True, padx=16, pady=8)
        lb = tk.Listbox(wrap, font=(FONT, 10), activestyle="none",
                        bg=COL["bg"], fg=COL["text"], bd=0, highlightthickness=1,
                        highlightbackground=COL["border"],
                        selectbackground=COL["sel"], selectforeground="white")
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb = ttk.Scrollbar(wrap, orient="vertical", command=lb.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        lb.configure(yscrollcommand=sb.set)

        def refresh():
            lb.delete(0, tk.END)
            files = sorted(p.name for p in doc_dir.iterdir() if p.is_file())
            for f in files:
                lb.insert(tk.END, f)
            count_lbl.config(text=f"{len(files)} document(s) — {doc_dir}")

        def add_docs():
            paths = filedialog.askopenfilenames(
                title="زيد وثائق", parent=win,
                filetypes=[("Tous", "*.*")])
            for src in paths:
                try:
                    shutil.copy2(src, doc_dir / Path(src).name)
                except OSError as exc:
                    messagebox.showerror("خطأ", str(exc), parent=win)
            refresh()

        def open_sel(_e=None):
            sel = lb.curselection()
            if not sel:
                return
            target = doc_dir / lb.get(sel[0])
            try:
                os.startfile(target)
            except (AttributeError, OSError):
                webbrowser.open(target.as_uri())

        def del_sel():
            sel = lb.curselection()
            if not sel:
                return
            name = lb.get(sel[0])
            if messagebox.askyesno("مسح", f"تمسح: {name} ؟", parent=win):
                try:
                    (doc_dir / name).unlink()
                except OSError as exc:
                    messagebox.showerror("خطأ", str(exc), parent=win)
                refresh()

        lb.bind("<Double-1>", open_sel)
        bb = tk.Frame(win, bg=COL["surface"])
        bb.pack(fill=tk.X, padx=16, pady=(0, 14))
        tb.Button(bb, text="➕  Ajouter", bootstyle="success",
                   command=add_docs).pack(side=tk.LEFT)
        tb.Button(bb, text="👁  Ouvrir", bootstyle="secondary-outline",
                   command=open_sel).pack(side=tk.LEFT, padx=6)
        tb.Button(bb, text="🗑  Supprimer", bootstyle="danger",
                   command=del_sel).pack(side=tk.LEFT)
        tb.Button(bb, text="📁  Dossier", bootstyle="secondary-outline",
                   command=lambda: os.startfile(doc_dir)).pack(side=tk.RIGHT)
        refresh()


    def _next_doc_ref(self, prefix: str) -> str:
        year = dt.date.today().year
        key = f"{prefix}-{year}"
        counters = self.config_data.setdefault("doc_counters", {})
        n = counters.get(key, 0) + 1
        counters[key] = n
        save_config(self.config_data)
        return f"{key}-{n:04d}"

    def _registry_path(self) -> Path:
        return self.store.path.parent / "documents_registry.json"

    def _register_doc(self, ref, dtype, rec, file):
        p = self._registry_path()
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
        except (OSError, ValueError):
            data = []
        data.append({
            "ref": ref, "type": dtype, "emp_id": rec.get("id", ""),
            "nom": rec.get("nom", ""), "date": dt.date.today().isoformat(),
            "file": str(file),
        })
        try:
            p.write_text(json.dumps(data, ensure_ascii=False, indent=2),
                         encoding="utf-8")
        except OSError:
            log.exception("Echec de l'ecriture du registre des documents")

    def open_registry(self):
        try:
            data = json.loads(self._registry_path().read_text(encoding="utf-8"))
        except (OSError, ValueError):
            data = []
        win = tk.Toplevel(self)
        win.title("Registre des documents")
        win.configure(bg=COL["surface"])
        win.geometry("680x460")
        tk.Label(win, text="🗂  Registre des documents générés", bg=COL["surface"],
                 fg=COL["brand"], font=(FONT, 14, "bold")).pack(
                     anchor=tk.W, padx=16, pady=(14, 6))
        cols = ("ref", "type", "nom", "date")
        tree = ttk.Treeview(win, columns=cols, show="headings")
        for c, lab, w in (("ref", "Référence", 150), ("type", "Type", 150),
                          ("nom", "Employé", 180), ("date", "Date", 100)):
            tree.heading(c, text=lab)
            tree.column(c, width=w, anchor=tk.W)
        tree.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 8))
        files = {}
        for i, d in enumerate(reversed(data)):
            iid = str(i)
            files[iid] = d.get("file", "")
            tree.insert("", tk.END, iid=iid, values=(
                d.get("ref", ""), d.get("type", ""), d.get("nom", ""),
                d.get("date", "")))

        def open_file(_e=None):
            sel = tree.selection()
            if not sel:
                return
            f = files.get(sel[0], "")
            if f and os.path.exists(f):
                try:
                    os.startfile(f)
                except (AttributeError, OSError):
                    webbrowser.open(Path(f).as_uri())
            else:
                messagebox.showinfo("مكاين", "الملف ماكاينش (تمسح أو تنقل).",
                                    parent=win)

        tree.bind("<Double-1>", open_file)
        bb = tk.Frame(win, bg=COL["surface"])
        bb.pack(fill=tk.X, padx=16, pady=(0, 14))
        tk.Label(bb, text=f"{len(data)} document(s)", bg=COL["surface"],
                 fg=COL["muted"], font=(FONT, 9)).pack(side=tk.LEFT)
        tb.Button(bb, text="👁  Ouvrir", bootstyle="success",
                   command=open_file).pack(side=tk.RIGHT)


    def generate_attestation(self, kind="travail"):
        if not HAS_FPDF:
            messagebox.showinfo("fpdf2 ناقصة", "ثبت:  pip install fpdf2")
            return
        if self.current_index is None:
            messagebox.showinfo("اختار خدام", "اختار شي خدام الأول.")
            return
        rec = self.records[self.current_index]
        if not rec.get("nom"):
            return
        p = compute_payroll(rec)
        nom = rec.get("nom", "")
        cin = rec.get("cin", "")
        poste = rec.get("poste", "")
        de = rec.get("date_embauche", "")
        contrat = rec.get("type_contrat", "")

        if kind == "salaire":
            titre = "ATTESTATION DE SALAIRE"
            corps = (
                f"Nous soussignés, {COMPANY_NAME}, attestons que "
                f"M./Mme {nom}, titulaire de la C.I.N. n° {cin}, est employé(e) "
                f"au sein de notre société en qualité de {poste} "
                f"depuis le {de}.\n\n"
                f"L'intéressé(e) perçoit une rémunération mensuelle de "
                f"{fmt_money(p['brut'])} brut, soit {fmt_money(p['net'])} net.\n\n"
                f"La présente attestation est délivrée à l'intéressé(e) pour "
                f"servir et valoir ce que de droit.")
        else:
            titre = "ATTESTATION DE TRAVAIL"
            corps = (
                f"Nous soussignés, {COMPANY_NAME}, attestons que "
                f"M./Mme {nom}, titulaire de la C.I.N. n° {cin}, "
                f"est employé(e) au sein de notre société en qualité de "
                f"{poste} depuis le {de}, dans le cadre d'un contrat "
                f"{contrat}.\n\n"
                f"La présente attestation est délivrée à l'intéressé(e) pour "
                f"servir et valoir ce que de droit.")

        pdf = FPDF(format="A4")
        pdf.set_auto_page_break(True, 20)
        font = self._setup_pdf_font(pdf)
        pdf.add_page()

        def t(s):
            return s if font == "uni" else \
                str(s).encode("latin-1", "replace").decode("latin-1")

        pdf.set_fill_color(15, 61, 46)
        pdf.rect(0, 0, 210, 26, "F")
        if LOGO_PATH and os.path.exists(LOGO_PATH):
            try:
                pdf.image(LOGO_PATH, x=12, y=4, h=18)
            except Exception:
                pass
        pdf.set_text_color(255, 255, 255)
        pdf.set_font(font, "B", 15)
        pdf.set_xy(0, 9)
        pdf.cell(210, 8, t(COMPANY_NAME), align="C")

        ref = self._next_doc_ref("ATT")
        pdf.set_text_color(100, 116, 139)
        pdf.set_font(font, "", 10)
        pdf.set_xy(20, 32)
        pdf.cell(170, 6, t(f"Réf : {ref}"), align="R")

        pdf.set_text_color(30, 41, 59)
        pdf.set_xy(0, 44)
        pdf.set_font(font, "B", 16)
        pdf.cell(210, 10, t(titre), align="C")
        pdf.ln(20)

        pdf.set_font(font, "", 12)
        pdf.set_x(20)
        pdf.multi_cell(170, 8, t(corps))
        pdf.ln(16)

        pdf.set_font(font, "", 11)
        ville_date = f"Fait le {dt.date.today().isoformat()}"
        pdf.set_x(20)
        pdf.cell(170, 8, t(ville_date), align="R")
        pdf.ln(16)
        pdf.set_x(20)
        pdf.cell(170, 8, t("Signature et cachet"), align="R")
        if SIGN_PATH and os.path.exists(SIGN_PATH):
            try:
                pdf.image(SIGN_PATH, x=135, y=pdf.get_y() + 4, w=50)
            except Exception:
                pass

        out_dir = self.store.path.parent / "fiches"
        out_dir.mkdir(exist_ok=True)
        safe = "".join(c for c in nom if c.isalnum() or c in " _-").strip()\
            .replace(" ", "_")
        out_file = out_dir / f"attestation_{kind}_{rec.get('id','')}_{safe}.pdf"
        try:
            pdf.output(str(out_file))
        except Exception as exc:
            messagebox.showerror("خطأ PDF", str(exc))
            return
        self._register_doc(ref, f"Attestation {kind}", rec, out_file)
        try:
            os.startfile(out_file)
        except (AttributeError, OSError):
            webbrowser.open(out_file.as_uri())
        self.set_status(f"الشهادة تصدّرات: {out_file.name}")


    def generate_contract(self):
        if not HAS_FPDF:
            messagebox.showinfo("fpdf2 ناقصة", "ثبت:  pip install fpdf2")
            return
        if self.current_index is None:
            messagebox.showinfo("اختار خدام", "اختار شي خدام الأول.")
            return
        rec = self.records[self.current_index]
        if not rec.get("nom"):
            return
        p = compute_payroll(rec)
        nom = rec.get("nom", "")
        contrat = (rec.get("type_contrat", "") or "CDI").upper()
        cdd = "CDD" in contrat
        titre = ("CONTRAT DE TRAVAIL À DURÉE DÉTERMINÉE" if cdd
                 else "CONTRAT DE TRAVAIL À DURÉE INDÉTERMINÉE")
        ref = self._next_doc_ref("CTR")

        if cdd and rec.get("date_fin_contrat"):
            duree = (f"Le présent contrat est conclu pour une durée déterminée, "
                     f"du {rec.get('date_embauche', '')} au "
                     f"{rec.get('date_fin_contrat', '')}.")
        elif cdd:
            duree = ("Le présent contrat est conclu pour une durée déterminée à "
                     "compter de la date d'embauche.")
        else:
            duree = ("Le présent contrat est conclu pour une durée indéterminée "
                     f"à compter du {rec.get('date_embauche', '')}.")

        articles = [
            ("Article 1 — Engagement",
             f"L'Employeur engage le Salarié en qualité de "
             f"{rec.get('poste', '')}, à compter du "
             f"{rec.get('date_embauche', '')}."),
            ("Article 2 — Durée du contrat", duree),
            ("Article 3 — Période d'essai",
             "Le présent contrat est assorti d'une période d'essai conforme à "
             "la législation du travail en vigueur (Code du travail marocain)."),
            ("Article 4 — Rémunération",
             f"En contrepartie de ses services, le Salarié percevra une "
             f"rémunération mensuelle brute de {fmt_money(p['brut'])}, "
             f"soumise aux cotisations légales (CNSS, AMO, IR)."),
            ("Article 5 — Lieu et horaires de travail",
             "Le Salarié exercera ses fonctions au siège de la société, "
             "selon les horaires en vigueur dans l'entreprise."),
            ("Article 6 — Obligations",
             "Le Salarié s'engage à respecter le règlement intérieur et à "
             "exécuter ses fonctions avec diligence. L'Employeur s'engage à "
             "déclarer le Salarié à la CNSS."),
        ]

        pdf = FPDF(format="A4")
        pdf.set_auto_page_break(True, 18)
        font = self._setup_pdf_font(pdf)
        pdf.add_page()

        def t(s):
            return s if font == "uni" else \
                str(s).encode("latin-1", "replace").decode("latin-1")

        pdf.set_fill_color(15, 61, 46)
        pdf.rect(0, 0, 210, 24, "F")
        if LOGO_PATH and os.path.exists(LOGO_PATH):
            try:
                pdf.image(LOGO_PATH, x=12, y=3, h=18)
            except Exception:
                pass
        pdf.set_text_color(255, 255, 255)
        pdf.set_font(font, "B", 14)
        pdf.set_xy(0, 8)
        pdf.cell(210, 8, t(COMPANY_NAME), align="C")

        pdf.set_text_color(100, 116, 139)
        pdf.set_font(font, "", 9)
        pdf.set_xy(20, 28)
        pdf.cell(170, 5, t(f"Réf : {ref}"), align="R")

        pdf.set_text_color(30, 41, 59)
        pdf.set_xy(0, 36)
        pdf.set_font(font, "B", 14)
        pdf.cell(210, 8, t(titre), align="C")
        pdf.ln(16)

        pdf.set_font(font, "", 11)
        pdf.set_x(18)
        pdf.multi_cell(174, 7, t(
            f"Entre les soussignés :\n"
            f"{COMPANY_NAME}, ci-après désignée « l'Employeur », d'une part,\n"
            f"Et M./Mme {nom}, titulaire de la C.I.N. n° {rec.get('cin', '')}, "
            f"demeurant à {rec.get('adresse', '') or '...'}, "
            f"ci-après désigné(e) « le Salarié », d'autre part.\n\n"
            f"Il a été convenu et arrêté ce qui suit :"))
        pdf.ln(2)

        for atitre, atexte in articles:
            pdf.set_font(font, "B", 11)
            pdf.set_x(18)
            pdf.multi_cell(174, 7, t(atitre))
            pdf.set_font(font, "", 11)
            pdf.set_x(18)
            pdf.multi_cell(174, 6, t(atexte))
            pdf.ln(1)

        pdf.ln(6)
        pdf.set_font(font, "", 11)
        pdf.set_x(18)
        pdf.cell(174, 6, t(f"Fait à ………………, le {dt.date.today().isoformat()}"),
                 align="R")
        pdf.ln(16)
        pdf.set_x(18)
        pdf.cell(87, 6, t("L'Employeur"), align="C")
        pdf.cell(87, 6, t("Le Salarié"), align="C")
        if SIGN_PATH and os.path.exists(SIGN_PATH):
            try:
                pdf.image(SIGN_PATH, x=40, y=pdf.get_y() + 4, w=45)
            except Exception:
                pass

        out_dir = self.store.path.parent / "fiches"
        out_dir.mkdir(exist_ok=True)
        safe = "".join(c for c in nom if c.isalnum() or c in " _-").strip()\
            .replace(" ", "_")
        out_file = out_dir / f"contrat_{rec.get('id', '')}_{safe}.pdf"
        try:
            pdf.output(str(out_file))
        except Exception as exc:
            messagebox.showerror("خطأ PDF", str(exc))
            return
        self._register_doc(ref, f"Contrat {contrat}", rec, out_file)
        try:
            os.startfile(out_file)
        except (AttributeError, OSError):
            webbrowser.open(out_file.as_uri())
        self.set_status(f"العقد تصدّر: {out_file.name}")


    def open_simulateur(self):
        win = tk.Toplevel(self)
        win.title("Simulateur de salaire")
        win.configure(bg=COL["surface"])
        win.geometry("420x440")
        tk.Label(win, text="🧮  Simulateur de salaire net", bg=COL["surface"],
                 fg=COL["brand"], font=(FONT, 14, "bold")).pack(
                     anchor=tk.W, padx=16, pady=(14, 8))

        form = tk.Frame(win, bg=COL["surface"])
        form.pack(fill=tk.X, padx=16)
        fields = [("Salaire de base", "salaire_base"), ("Primes", "primes"),
                  ("Personnes à charge", "personnes_charge")]
        sim_vars = {}
        for i, (label, key) in enumerate(fields):
            tk.Label(form, text=label, bg=COL["surface"], fg=COL["text"],
                     font=(FONT, 10)).grid(row=i, column=0, sticky=tk.W, pady=5)
            v = tk.StringVar(value="0")
            sim_vars[key] = v
            ttk.Entry(form, textvariable=v, width=16).grid(
                row=i, column=1, sticky=tk.E, pady=5, padx=(8, 0))
        form.columnconfigure(1, weight=1)

        res = tk.Frame(win, bg=COL["surface"])
        res.pack(fill=tk.BOTH, expand=True, padx=16, pady=12)
        res_vars = {}

        def add_res(key, label, strong=False):
            line = tk.Frame(res, bg=COL["surface"])
            line.pack(fill=tk.X, pady=2)
            f = (FONT, 12, "bold") if strong else (FONT, 10)
            tk.Label(line, text=label, bg=COL["surface"],
                     fg=COL["brand2"] if strong else COL["muted"],
                     font=f).pack(side=tk.LEFT)
            var = tk.StringVar(value="—")
            res_vars[key] = var
            tk.Label(line, textvariable=var, bg=COL["surface"],
                     fg=COL["brand2"] if strong else COL["text"],
                     font=f).pack(side=tk.RIGHT)

        for key, label in [("brut", "Salaire brut"), ("cnss", "CNSS"),
                           ("amo", "AMO"), ("ir", "IR"),
                           ("total_retenues", "Total retenues")]:
            add_res(key, label)
        tk.Frame(res, bg=COL["border"], height=1).pack(fill=tk.X, pady=6)
        add_res("net", "Salaire NET", strong=True)

        def recompute(*_):
            rec = {k: v.get() for k, v in sim_vars.items()}
            pr = compute_payroll(rec)
            for key, var in res_vars.items():
                var.set(fmt_money(pr[key]))

        for v in sim_vars.values():
            v.trace_add("write", recompute)
        recompute()


    def open_folder(self):
        folder = self.store.path.parent
        try:
            os.startfile(folder)
        except AttributeError:
            webbrowser.open(folder.as_uri())

    def set_status(self, text: str):
        self.status_var.set(text)

    def toast(self, message: str, kind: str = "success"):
        self.set_status(message)
        if HAS_TB_EXTRAS:
            try:
                ToastNotification(title="Gestion des Employés", message=message,
                                  duration=2600, bootstyle=kind,
                                  position=(20, 60, "se")).show_toast()
            except Exception:  # noqa: BLE001
                pass

    def _focus_search(self):
        if hasattr(self, "search_entry"):
            self.search_entry.focus_set()

    def _bind_shortcuts(self):
        self.bind("<Control-n>", lambda e: self.new_record())
        self.bind("<Control-s>", lambda e: self.save_record())
        self.bind("<Control-f>", lambda e: self._focus_search())
        self.bind("<Control-p>", lambda e: self.export_pdf())

    def _show_splash(self):
        try:
            sp = tk.Toplevel(self)
            sp.overrideredirect(True)
            w, h = 420, 240
            x = (self.winfo_screenwidth() - w) // 2
            y = (self.winfo_screenheight() - h) // 2
            sp.geometry(f"{w}x{h}+{x}+{y}")
            frame = tk.Frame(sp, bg=COL["brand"])
            frame.pack(fill=tk.BOTH, expand=True)
            tk.Label(frame, text="👥", bg=COL["brand"], fg="white",
                     font=(FONT, 44)).pack(pady=(34, 4))
            tk.Label(frame, text="Gestion des Employés & Paie", bg=COL["brand"],
                     fg="white", font=(FONT, 15, "bold")).pack()
            tk.Label(frame, text=f"Version {APP_VERSION}", bg=COL["brand"],
                     fg="#a7d7bd", font=(FONT, 9)).pack(pady=(2, 14))
            pb = ttk.Progressbar(frame, mode="indeterminate", length=240)
            pb.pack()
            pb.start(12)
            self._splash = sp
            sp.update()
        except Exception:  # noqa: BLE001
            self._splash = None

    def _close_splash(self):
        sp = getattr(self, "_splash", None)
        if sp is not None:
            try:
                sp.destroy()
            except Exception:  # noqa: BLE001
                pass
            self._splash = None
        self.deiconify()


def main():
    app = EmployeeApp()
    app.mainloop()


if __name__ == "__main__":
    main()
